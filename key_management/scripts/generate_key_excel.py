"""
Key Management Excel Report Generator

Generates an editable Excel file tracking which employees hold physical key sets
for the 4 déchetteries (Pépinière, St Germain, Polignac, Sanssac).
Each key set contains Portail + Local keys attached together.

Data sources:
  - data/employees.json: Employee records with category and déchetterie
  - data/keys.json: Number of key sets per déchetterie
  - data/assignments.json: Current key set assignments to employees
"""

import json
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Protection
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows


def load_json(filepath):
    """Load JSON file from data directory."""
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_new_employees_from_excel(excel_path):
    """Read new employees from the 'Nouvel Employé' sheet in the previous Excel report."""
    if not excel_path.exists():
        return []
    
    try:
        wb = load_workbook(excel_path)
        
        # Try to find the "Nouvel Employé" sheet
        if "Nouvel Employé" not in wb.sheetnames:
            wb.close()
            return []
        
        ws = wb["Nouvel Employé"]
        
        new_employees = []
        # Read rows starting from row 6 (after headers)
        for row in range(6, ws.max_row + 1):
            name_cell = ws.cell(row=row, column=1).value
            category_cell = ws.cell(row=row, column=2).value
            
            # If both name and category are filled, it's a new employee
            if name_cell and category_cell and isinstance(name_cell, str) and isinstance(category_cell, str):
                name = name_cell.strip()
                category = category_cell.strip()
                
                # Skip empty entries
                if name and category:
                    new_employees.append({"name": name, "category": category})
        
        wb.close()
        return new_employees
    except Exception as e:
        print(f"⚠ Could not read new employees from Excel: {e}")
        return []


def get_key_assignments_from_excel(excel_path, sorted_déchetteries):
    """Read existing key assignments from the previous Excel report."""
    if not excel_path.exists():
        return {}
    
    try:
        wb = load_workbook(excel_path)
        
        if "Gestion des Clés" not in wb.sheetnames:
            wb.close()
            return {}
        
        ws = wb["Gestion des Clés"]
        
        assignments = {}  # (employee_name, déchetterie) -> count
        
        # Find header row (row 5)
        header_row = 5
        
        # Get column indices for each déchetterie
        déchetterie_cols = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col_idx).value
            if header in sorted_déchetteries:
                déchetterie_cols[header] = col_idx
        
        # Read employee rows (starting from row 6, skip summary rows)
        for row in range(6, ws.max_row - 5):  # -5 to skip TOTAL, DISPONIBLES, TOTAL CLÉS rows
            name_cell = ws.cell(row=row, column=1).value
            
            # Skip empty rows and section headers
            if not name_cell or not isinstance(name_cell, str):
                continue
            
            name = name_cell.strip()
            if not name or name.upper() in ["EMPLOYÉ", "UTILISÉES", "DISPONIBLES", "TOTAL CLÉS"]:
                continue
            
            # Read key counts for each déchetterie
            for déchetterie, col_idx in déchetterie_cols.items():
                value = ws.cell(row=row, column=col_idx).value
                
                # Convert to int if numeric, otherwise 0
                try:
                    count = int(value) if value else 0
                    if count > 0:  # Only store non-zero values
                        assignments[(name, déchetterie)] = count
                except (ValueError, TypeError):
                    pass
        
        wb.close()
        return assignments
    except Exception as e:
        print(f"⚠ Could not read key assignments from Excel: {e}")
        return {}


def get_employee_info(employees, name):
    """Find employee info by name. Returns dict or None."""
    if not name or name.strip() == "":
        return None
    for emp in employees:
        if emp['name'].lower() == name.lower():
            return emp
    return None


def build_employee_lookup(employees):
    """Build a lookup dict of employee names (lowercase) -> full info."""
    lookup = {}
    for emp in employees:
        lookup[emp['name'].lower()] = emp
    return lookup


def generate_key_excel():
    """Generate Excel report with all key assignments in one comprehensive sheet."""
    
    # Get script directory to load relative data files
    script_dir = Path(__file__).parent
    data_dir = script_dir.parent / 'data'
    output_dir = script_dir.parent / 'output'
    
    # Create output directory if it doesn't exist
    output_dir.mkdir(exist_ok=True)
    
    # Load data
    employees = load_json(data_dir / 'employees.json')
    keys_config = load_json(data_dir / 'keys.json')
    
    # Get sorted list of déchetteries (needed for reading assignments)
    sorted_déchetteries = sorted(keys_config.keys())
    
    # Read new employees from previous Excel (if it exists)
    prev_excel = output_dir / 'key_report_latest.xlsx'
    new_employees_from_excel = get_new_employees_from_excel(prev_excel)
    
    # Merge new employees from Excel with base employees
    # Avoid duplicates by checking names
    existing_names = {emp['name'].lower() for emp in employees}
    for new_emp in new_employees_from_excel:
        if new_emp['name'].lower() not in existing_names:
            employees.append(new_emp)
            existing_names.add(new_emp['name'].lower())
    
    # Read key assignments from previous Excel (if it exists)
    assignment_lookup = get_key_assignments_from_excel(prev_excel, sorted_déchetteries)
    
    # Create workbook with single sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Gestion des Clés"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    assigned_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    pale_yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border_top = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='thin')
    )
    
    thick_border_bottom = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='thin'),
        bottom=Side(style='medium')
    )
    
    thick_border_middle = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Define category ranking (higher rank = priority in sorting)
    rank_order = {'STAFF': 1, 'Staff': 1, 'CIP': 2, 'CDI': 3, 'CDDI': 4}
    
    # Get employees sorted by rank then name (from employees.json)
    employees_with_keys = []
    for emp in employees:
        emp_name = emp['name']
        category = emp['category']
        rank = rank_order.get(category, 999)
        employees_with_keys.append((emp_name, emp, rank))
    
    employees_with_keys.sort(key=lambda x: (x[2], x[0]))  # Sort by rank, then name
    
    # Calculate last column letter for full width merges
    num_cols = len(sorted_déchetteries) + 3  # Name + Rank + Déchetteries + Total
    last_col_letter = chr(64 + num_cols)
    
    # Write title - merge across all columns
    ws['A1'] = "GESTION DES CLÉS - RAPPORT COMPLET"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws.row_dimensions[1].height = 25
    
    ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10, color="FFFFFF")
    ws['A2'].fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws.row_dimensions[2].height = 18
    
    # Write table section title
    ws['A4'] = "ASSIGNATION DES CLÉS PAR EMPLOYÉ"
    ws['A4'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A4'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A4:{last_col_letter}4')
    ws.row_dimensions[4].height = 22
    
    # Table headers: Employé, Rang, [Déchetteries...]
    header_cols = ["Employé", "Rang"] + sorted_déchetteries + ["Total"]
    for col_idx, header in enumerate(header_cols, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    # Define separator row fill color (more visible)
    separator_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Add employee rows
    data_row = 6
    first_data_row = 6
    previous_category = None
    rows_by_category = {}  # Track rows for each category
    current_category_rows = []
    
    for emp_idx, (emp_name, emp_info, rank) in enumerate(employees_with_keys):
        current_category = emp_info['category']
        
        # If category changed, start new group
        if previous_category is not None and previous_category != current_category:
            rows_by_category[previous_category] = current_category_rows
            current_category_rows = []
            
            # Add separator row with NO BORDER
            separator_font = Font(italic=True, color="666666", size=9)
            for col in range(1, len(header_cols) + 1):
                cell = ws.cell(row=data_row, column=col)
                cell.border = Border()  # NO BORDER
                cell.protection = Protection(locked=True)
                cell.font = separator_font
                cell.alignment = left_align if col == 1 else center_align
            data_row += 1
        
        # Store this row for the current category
        current_category_rows.append(data_row)
        
        # Employee name
        ws.cell(row=data_row, column=1).value = emp_name
        ws.cell(row=data_row, column=1).alignment = left_align
        ws.cell(row=data_row, column=1).font = Font(bold=True)
        ws.cell(row=data_row, column=1).fill = pale_yellow_fill
        
        # Rank
        ws.cell(row=data_row, column=2).value = emp_info['category']
        ws.cell(row=data_row, column=2).alignment = center_align
        ws.cell(row=data_row, column=2).fill = pale_yellow_fill
        
        # Keys per déchetterie - editable cells
        for col_idx, déchetterie in enumerate(sorted_déchetteries, start=3):
            key_count = assignment_lookup.get((emp_name, déchetterie), 0)
            
            cell = ws.cell(row=data_row, column=col_idx)
            cell.value = key_count if key_count > 0 else None
            cell.alignment = center_align
            if key_count > 0:
                cell.fill = assigned_fill
                cell.font = Font(bold=True)
            cell.protection = Protection(locked=False)
        
        # Total column with formula
        total_col = len(header_cols)
        first_déchetterie_col = chr(65 + 2)  # Column C
        last_déchetterie_col = chr(65 + len(header_cols) - 2)  # Last déchetterie column
        total_formula = f"=SUM({first_déchetterie_col}{data_row}:{last_déchetterie_col}{data_row})"
        
        cell = ws.cell(row=data_row, column=total_col)
        cell.value = total_formula
        cell.alignment = center_align
        cell.fill = assigned_fill
        cell.font = Font(bold=True)
        
        previous_category = current_category
        data_row += 1
    
    # Store last category
    if employees_with_keys:
        rows_by_category[current_category] = current_category_rows
    
    # Apply borders to each category group
    for category, category_rows in rows_by_category.items():
        if not category_rows:
            continue
        
        first_row = category_rows[0]
        last_row = category_rows[-1]
        
        for row in category_rows:
            for col in range(1, len(header_cols) + 1):
                cell = ws.cell(row=row, column=col)
                
                # Determine which border style to use
                is_first = (row == first_row)
                is_last = (row == last_row)
                
                if is_first and is_last:
                    # Single row category
                    border_style = Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium')
                    )
                elif is_first:
                    # First row of group
                    border_style = thick_border_top
                elif is_last:
                    # Last row of group
                    border_style = thick_border_bottom
                else:
                    # Middle rows
                    border_style = thick_border_middle
                
                cell.border = border_style
    
    last_data_row = data_row - 1
    
    # Create data validation with numeric values (0-5)
    dv = DataValidation(type="list", formula1='"0,1,2,3,4,5"', allow_blank=True)
    dv.error = 'Sélectionnez 0 à 5'
    dv.errorTitle = 'Valeur invalide'
    dv.prompt = 'Nombre de clés (0-5)'
    dv.promptTitle = 'Nombre de clés'
    ws.add_data_validation(dv)
    
    # Add data validation to data cells
    for col_idx, déchetterie in enumerate(sorted_déchetteries, start=3):
        col_letter = chr(65 + col_idx - 1)
        
        # Add data validation to each cell in the range
        for row in range(first_data_row, last_data_row + 1):
            cell_ref = f"{col_letter}{row}"
            dv.add(cell_ref)
    
    # Add totals row
    total_row = data_row + 1
    ws.cell(row=total_row, column=1).value = "UTILISÉES"
    ws.cell(row=total_row, column=1).border = border
    ws.cell(row=total_row, column=1).alignment = center_align
    ws.cell(row=total_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    
    ws.cell(row=total_row, column=2).border = border
    ws.cell(row=total_row, column=2).fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    
    # Totals for each déchetterie with formulas
    for col_idx, déchetterie in enumerate(sorted_déchetteries, start=3):
        col_letter = chr(65 + col_idx - 1)
        total_formula = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        
        cell = ws.cell(row=total_row, column=col_idx)
        cell.value = total_formula
        cell.border = border
        cell.alignment = center_align
        cell.fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Grand total formula
    first_déchetterie_col = chr(65 + 2)
    last_déchetterie_col = chr(65 + len(header_cols) - 2)
    grand_total_formula = f"=SUM({first_déchetterie_col}{total_row}:{last_déchetterie_col}{total_row})"
    
    cell = ws.cell(row=total_row, column=len(header_cols))
    cell.value = grand_total_formula
    cell.border = border
    cell.alignment = center_align
    cell.fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    
    # Add available keys row
    available_row = total_row + 1
    ws.cell(row=available_row, column=1).value = "DISPONIBLES"
    ws.cell(row=available_row, column=1).border = border
    ws.cell(row=available_row, column=1).alignment = center_align
    ws.cell(row=available_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=available_row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    ws.cell(row=available_row, column=2).border = border
    ws.cell(row=available_row, column=2).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Available keys for each déchetterie (Total - Assigned)
    for col_idx, déchetterie in enumerate(sorted_déchetteries, start=3):
        col_letter = chr(65 + col_idx - 1)
        # Available = Total keys (from total_keys_row) - Assigned keys (from total_row)
        # Reference will be updated after creating the total_keys_row
        
        cell = ws.cell(row=available_row, column=col_idx)
        cell.border = border
        cell.alignment = center_align
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        # Placeholder - will set formula after creating total_keys_row
        cell.value = None
    
    # Grand total available formula (placeholder)
    cell = ws.cell(row=available_row, column=len(header_cols))
    cell.border = border
    cell.alignment = center_align
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.value = None
    
    # Add total keys row (editable)
    total_keys_row = available_row + 1
    ws.cell(row=total_keys_row, column=1).value = "TOTAL CLÉS"
    ws.cell(row=total_keys_row, column=1).border = border
    ws.cell(row=total_keys_row, column=1).alignment = center_align
    ws.cell(row=total_keys_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=total_keys_row, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    ws.cell(row=total_keys_row, column=2).border = border
    ws.cell(row=total_keys_row, column=2).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Editable total keys for each déchetterie
    for col_idx, déchetterie in enumerate(sorted_déchetteries, start=3):
        col_letter = chr(65 + col_idx - 1)
        total_keys = keys_config[déchetterie]
        
        cell = ws.cell(row=total_keys_row, column=col_idx)
        cell.value = total_keys
        cell.border = border
        cell.alignment = center_align
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        # Make editable
        cell.protection = Protection(locked=False)
    
    # Grand total for total keys row
    grand_total_keys_formula = f"=SUM({first_déchetterie_col}{total_keys_row}:{last_déchetterie_col}{total_keys_row})"
    cell = ws.cell(row=total_keys_row, column=len(header_cols))
    cell.value = grand_total_keys_formula
    cell.border = border
    cell.alignment = center_align
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    
    # Now set the available keys formulas (referencing the editable total_keys_row)
    for col_idx, déchetterie in enumerate(sorted_déchetteries, start=3):
        col_letter = chr(65 + col_idx - 1)
        # Available = Total keys (row) - Assigned keys (total_row)
        available_formula = f"={col_letter}{total_keys_row}-{col_letter}{total_row}"
        
        cell = ws.cell(row=available_row, column=col_idx)
        cell.value = available_formula
    
    # Grand total available formula
    grand_available_formula = f"=SUM({first_déchetterie_col}{available_row}:{last_déchetterie_col}{available_row})"
    cell = ws.cell(row=available_row, column=len(header_cols))
    cell.value = grand_available_formula
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 10
    for col_idx, _ in enumerate(sorted_déchetteries, start=3):
        ws.column_dimensions[chr(64 + col_idx)].width = 15
    ws.column_dimensions[chr(64 + len(header_cols))].width = 10
    
    # Protect sheet but allow editing cells
    ws.protection.sheet = True
    ws.protection.enable()
    
    # Create second sheet for new employees
    ws_new_emp = wb.create_sheet(title="Nouvel Employé")
    
    # Title
    ws_new_emp['A1'] = "AJOUTER NOUVEL EMPLOYÉ"
    ws_new_emp['A1'].font = Font(bold=True, size=13)
    ws_new_emp.merge_cells('A1:B1')
    
    # Instructions
    ws_new_emp['A3'] = "Entrez les informations du nouvel employé ci-dessous:"
    ws_new_emp['A3'].font = Font(italic=True, size=10)
    ws_new_emp.merge_cells('A3:B3')
    
    # Headers
    ws_new_emp['A5'] = "Nom"
    ws_new_emp['B5'] = "Catégorie"
    for col in ['A', 'B']:
        cell = ws_new_emp[f'{col}5']
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    # Create data validation for category dropdown
    category_dv = DataValidation(type="list", formula1='"STAFF,CIP,CDI,CDDI"', allow_blank=True)
    category_dv.error = 'Sélectionnez une catégorie'
    category_dv.errorTitle = 'Catégorie invalide'
    category_dv.prompt = 'STAFF, CIP, CDI, ou CDDI'
    category_dv.promptTitle = 'Catégorie'
    ws_new_emp.add_data_validation(category_dv)
    
    # Add 10 empty rows for new employees
    for i in range(10):
        row_num = 6 + i
        for col in ['A', 'B']:
            cell = ws_new_emp[f'{col}{row_num}']
            cell.border = border
            cell.alignment = left_align
            cell.protection = Protection(locked=False)
        
        # Add category dropdown to column B
        category_dv.add(f'B{row_num}')
    
    ws_new_emp.column_dimensions['A'].width = 25
    ws_new_emp.column_dimensions['B'].width = 15
    
    # Protect new employees sheet
    ws_new_emp.protection.sheet = True
    ws_new_emp.protection.enable()
    
    # Update get_new_employees_from_excel to read from "Nouvel Employé" sheet
    
    # Save workbook
    timestamp = datetime.now().strftime('%Y-%m-%d')
    output_file = output_dir / f'key_report_{timestamp}.xlsx'
    wb.save(output_file)
    
    # Also save as "latest" for reading new employees on next generation
    latest_file = output_dir / 'key_report_latest.xlsx'
    wb.save(latest_file)
    
    print(f"✓ Excel report generated: {output_file}")
    return str(output_file)


if __name__ == '__main__':
    generate_key_excel()
