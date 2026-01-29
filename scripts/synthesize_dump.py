"""
Synthétise les données dump depuis la base de données en un fichier Excel au format COLLECTES.

Ce script lit les données dump depuis la base de données (même source que le backend)
et génère un fichier Excel synthétisé avec les mêmes totaux que le frontend et le backend.

Correction majeure: Utilise maintenant le mapping map_category_to_collectes() du backend
au lieu de chercher EVACUATION DECHETS manuellement, garantissant l'alignement des totaux.
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Import mapping functions from mappings.py
import sys
script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

from mappings import (
    DECHETTERIE_MAPPING, 
    map_category_to_collectes,
    CATEGORY_COLUMNS,
    FINAL_FLUXES
)


def _get_project_paths():
    """Get project root, input, and output directories."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    input_dir = os.path.join(project_root, 'input')
    output_dir = os.path.join(project_root, 'output')
    return project_root, input_dir, output_dir


def synthesize_dump(output_file, year=2025):
    """
    Synthétise les données dump depuis la base de données en un fichier Excel au format COLLECTES.
    
    Utilise le même mapping que le backend (dump_stats_service.py) pour garantir
    que les totaux du script correspondent aux totaux du frontend et du backend.
    
    Args:
        output_file: Path for output file
        year: Year for the dump (default: 2025)
    """
    
    print(f"\n{'='*70}")
    print(" " * 15 + "SYNTHESE DUMP - ALIGNEMENT BACKEND/FRONTEND")
    print(f"{'='*70}")
    
    # Import database functions
    current_file = Path(__file__).resolve()
    project_root = current_file.parent.parent
    server_dir = project_root / 'server'
    if str(server_dir) not in sys.path:
        sys.path.insert(0, str(server_dir))
    
    try:
        from services.db import get_dump_connection, init_dump_db
    except ImportError:
        print(f"\n[ERREUR] Impossible d'importer les fonctions de base de données")
        return None
    
    # Initialize database and read data
    print(f"\n[LECTURE] Lecture depuis la base de données dump-{year}.db...")
    try:
        init_dump_db(year)
        with get_dump_connection(year) as conn:
            df = pd.read_sql(
                """
                SELECT date, lieu_collecte, categorie, sous_categorie, flux, orientation, poids
                FROM raw_dump
                """,
                conn
            )
        print(f"   [OK] {len(df):,} enregistrements lus")
    except Exception as e:
        print(f"\n[ERREUR] Impossible de lire la base de données : {str(e)}")
        return None
    
    if df.empty:
        print(f"\n[ERREUR] Aucune donnée dans la base de données")
        return None
    
    # Prepare data (identical to backend logic)
    print(f"\n[TRAITEMENT] Traitement des données (même logique que le backend)...")
    
    # 1. Filter invalid dates
    df['Date'] = pd.to_datetime(df['date'], errors='coerce')
    total_brut = df['poids'].sum() / 1000
    df = df[df['Date'].notna()].copy()
    total_apres_date = df['poids'].sum() / 1000 if not df.empty else 0
    print(f"   Total brut: {total_brut:,.2f} tonnes")
    print(f"   Total après filtre dates invalides: {total_apres_date:,.2f} tonnes")
    
    if df.empty:
        print(f"\n[ERREUR] Aucune date valide dans les données")
        return None
    
    # 2. Map déchetteries (same logic as backend)
    df['Dechetterie'] = df['lieu_collecte'].map(DECHETTERIE_MAPPING)
    df['Dechetterie'] = df['Dechetterie'].fillna(df['lieu_collecte'])
    # Assign empty locations to Pépinière
    empty_mask = df['lieu_collecte'].isna() | (df['lieu_collecte'].astype(str).str.strip() == '')
    df.loc[empty_mask, 'Dechetterie'] = 'Pépinière'
    # Assign APPORT VOLONTAIRE/SUR SITE to Pépinière
    apport_mask = (
        df['lieu_collecte'].astype(str).str.upper().eq('APPORT VOLONTAIRE') |
        df['lieu_collecte'].astype(str).str.upper().eq('APPORT SUR SITE')
    )
    df.loc[apport_mask, 'Dechetterie'] = 'Pépinière'
    # Assign unmapped (NaN or 'nan' string) to Pépinière
    nan_mask = (
        df['Dechetterie'].isna() |
        (df['Dechetterie'].astype(str).str.upper().str.strip().isin(['NAN', '']))
    )
    df.loc[nan_mask, 'Dechetterie'] = 'Pépinière'
    
    # 3. Map categories using the SAME function as backend
    # This is the KEY FIX - we now use map_category_to_collectes() instead of manual search
    df['MappedCategory'] = df.apply(
        lambda row: map_category_to_collectes(
            row['categorie'],
            row['sous_categorie'],
            row['flux'],
            row['orientation']
        ),
        axis=1
    )
    
    # Fill unmapped categories with AUTRES
    df.loc[df['MappedCategory'].isna(), 'MappedCategory'] = 'AUTRES'
    
    # Log unmapped data
    non_mappees = df[df['MappedCategory'] == 'AUTRES']
    if len(non_mappees) > 0:
        total_autres = non_mappees['poids'].sum() / 1000
        print(f"   {len(non_mappees):,} lignes non mappées ({total_autres:,.2f} tonnes) → AUTRES")
    
    # 4. Extract month
    month_names_fr = {
        1: 'JANVIER', 2: 'FEVRIER', 3: 'MARS', 4: 'AVRIL',
        5: 'MAI', 6: 'JUIN', 7: 'JUILLET', 8: 'AOUT',
        9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DECEMBRE'
    }
    df['Month'] = df['Date'].dt.month
    df['MonthName'] = df['Month'].map(month_names_fr)
    df = df[df['MonthName'].notna()].copy()
    
    # Get date range
    date_start = df['Date'].min()
    date_end = df['Date'].max()
    date_range_str = f"{date_start.strftime('%d/%m/%Y')} - {date_end.strftime('%d/%m/%Y')}"
    
    # Get unique déchetteries
    unique_dechetteries = df['Dechetterie'].unique().tolist()
    
    print(f"   Période : {date_range_str}")
    print(f"   Déchetteries : {len(unique_dechetteries)}")
    print(f"   Catégories mappées : {df['MappedCategory'].nunique()}")
    
    # Order déchetteries (standard first, then alphabetical)
    standard_order = ['Pépinière', 'Sanssac', 'St Germain', 'Polignac', 'Yssingeaux', 'Bas-en-Basset', 'Monistrol']
    special_cases = [d for d in unique_dechetteries if d not in standard_order]
    ordered_dechetteries = [d for d in standard_order if d in unique_dechetteries] + sorted(special_cases)
    
    print(f"   Traitement de {len(ordered_dechetteries)} déchetteries...")
    
    # Get column lists from mappings.py (same as backend)
    category_columns = CATEGORY_COLUMNS.copy()
    final_fluxes = FINAL_FLUXES.copy()
    all_columns = category_columns + final_fluxes
    
    # Month order
    month_order = ['JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN',
                   'JUILLET', 'AOUT', 'SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DECEMBRE']
    
    # Build Excel structure
    all_rows = []
    row_positions = {}  # Track where each déchetterie's data is
    dechetterie_totals = {}  # Store totals by déchetterie for percentage calculations
    grand_totals = {col: 0 for col in all_columns}
    grand_totals['TOTAL'] = 0
    grand_totals['sans massicot et démantèlement'] = 0
    
    # Title row
    title_text = f"COLLECTES RECYCLERIE {date_range_str}"
    all_rows.append([title_text] + [''] * (len(all_columns) + 2))
    all_rows.append([''] * (len(all_columns) + 3))  # Empty row
    all_rows.append([''] * (len(all_columns) + 3))  # Empty row
    
    # Process each déchetterie
    for dech in ordered_dechetteries:
        dech_df = df[df['Dechetterie'] == dech].copy()
        
        if len(dech_df) == 0:
            continue
        
        # Initialize totals for this déchetterie
        dechetterie_totals[dech] = {col: 0 for col in all_columns}
        
        # Separate categories from final fluxes
        dech_df_categories = dech_df[~dech_df['MappedCategory'].isin(final_fluxes)].copy()
        dech_df_final_fluxes = dech_df[dech_df['MappedCategory'].isin(final_fluxes)].copy()
        
        # Pivot categories
        if not dech_df_categories.empty:
            summary_cat = dech_df_categories.groupby(['MonthName', 'MappedCategory'])['poids'].sum().reset_index()
            pivot_cat = summary_cat.pivot_table(
                index='MonthName',
                columns='MappedCategory',
                values='poids',
                aggfunc='sum',
                fill_value=0
            ).reset_index()
        else:
            pivot_cat = pd.DataFrame({'MonthName': month_order})
        
        # Pivot final fluxes
        if not dech_df_final_fluxes.empty:
            summary_flux = dech_df_final_fluxes.groupby(['MonthName', 'MappedCategory'])['poids'].sum().reset_index()
            pivot_flux = summary_flux.pivot_table(
                index='MonthName',
                columns='MappedCategory',
                values='poids',
                aggfunc='sum',
                fill_value=0
            ).reset_index()
        else:
            pivot_flux = pd.DataFrame({'MonthName': month_order})
        
        # Merge and reorder
        result_df = pivot_cat.merge(pivot_flux, on='MonthName', how='outer')
        ordered_df_temp = pd.DataFrame({'MonthName': month_order})
        result_df = ordered_df_temp.merge(result_df, on='MonthName', how='left').fillna(0)
        
        # Header row for this déchetterie
        row_positions[dech] = {'header_row': len(all_rows) + 1}
        header_row = [dech.upper()] + all_columns + ['TOTAL', 'Sans massicot et démantèlement']
        all_rows.append(header_row)
        
        # Data rows
        row_positions[dech]['data_start'] = len(all_rows) + 1
        for _, row in result_df.iterrows():
            data_row = [row['MonthName']]
            row_total = 0
            
            # Add all columns (categories + final fluxes)
            for col in all_columns:
                val = row.get(col, 0)
                data_row.append(val)
                if isinstance(val, (int, float)):
                    row_total += val
                    grand_totals[col] += val
                    dechetterie_totals[dech][col] += val
            
            # TOTAL column
            data_row.append(row_total)
            grand_totals['TOTAL'] += row_total
            
            # Sans massicot et démantèlement
            massicot = row.get('MASSICOT', 0) if 'MASSICOT' in row else 0
            demantelement = row.get('DEMANTELEMENT', 0) if 'DEMANTELEMENT' in row else 0
            sans_massicot = row_total - massicot - demantelement
            data_row.append(sans_massicot)
            grand_totals['sans massicot et démantèlement'] += sans_massicot
            dechetterie_totals[dech]['sans massicot et démantèlement'] = dechetterie_totals[dech].get('sans massicot et démantèlement', 0) + sans_massicot
            
            # Empty columns
            data_row.extend(['', '', ''])
            
            all_rows.append(data_row)
        
        row_positions[dech]['data_end'] = len(all_rows)
        
        # Total row for this déchetterie
        row_positions[dech]['total_row'] = len(all_rows) + 1
        total_row = ['Total'] + [''] * (len(all_columns) + 2)
        all_rows.append(total_row)
        
        # Empty row
        all_rows.append([''] * (len(all_columns) + 3))
    
    # Add some empty rows before grand total
    all_rows.append([''] * (len(all_columns) + 3))
    all_rows.append([''] * (len(all_columns) + 3))
    
    # Add categories reminder row (legend)
    categories_reminder_row_num = len(all_rows) + 1
    categories_reminder_row = ['CATEGORIES'] + all_columns + ['TOTAL', 'SANS MASSICOT\nET DEMANTELEMENT']
    all_rows.append(categories_reminder_row)
    
    # Grand total row
    grand_total_row_num = len(all_rows) + 1
    grand_total_row = ['TOTAL'] + [''] * (len(all_columns) + 1)
    all_rows.append(grand_total_row)
    
    # Add empty rows before percentages section
    all_rows.append([''] * (len(all_columns) + 3))
    all_rows.append([''] * (len(all_columns) + 3))
    all_rows.append([''] * (len(all_columns) + 3))
    
    # Add percentages section title
    percentages_title_row_num = len(all_rows) + 1
    percentages_title_row = ['POURCENTAGES PAR CATEGORIE ET DECHETTERIE'] + [''] * (len(all_columns) + 2)
    all_rows.append(percentages_title_row)
    
    all_rows.append([''] * (len(all_columns) + 3))  # Empty row
    
    # Build percentages table
    percentages_rows_start = len(all_rows) + 1
    percentages_header_row_num = percentages_rows_start
    percentages_header_row = ['DECHETTERIE'] + all_columns + ['TOTAL']
    all_rows.append(percentages_header_row)
    
    # Add percentage rows for each déchetterie
    percentages_data_rows = {}
    for dech in ordered_dechetteries:
        if dech not in row_positions:
            continue
        percentages_data_rows[dech] = len(all_rows) + 1
        pct_row = [dech.upper()] + [''] * (len(all_columns) + 1)
        all_rows.append(pct_row)
    
    # Print diagnostics
    print(f"\n[DIAGNOSTIC] Totaux avant écriture Excel:")
    print(f"   Grand Total: {grand_totals['TOTAL']/1000:,.2f} tonnes")
    print(f"   DECHETS ULTIMES: {grand_totals.get('DECHETS ULTIMES', 0)/1000:,.2f} tonnes")
    print(f"   MASSICOT: {grand_totals.get('MASSICOT', 0)/1000:,.2f} tonnes")
    print(f"   DEMANTELEMENT: {grand_totals.get('DEMANTELEMENT', 0)/1000:,.2f} tonnes")
    
    # Write to Excel
    print(f"\n[ECRITURE] Ecriture du fichier Excel...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        final_df = pd.DataFrame(all_rows)
        final_df.to_excel(writer, sheet_name='CALCUL POIDS', index=False, header=False)
    
    # Update formulas
    print("   [OK] Mise à jour des formules...")
    wb = load_workbook(output_file)
    ws = wb['CALCUL POIDS']
    
    # Update déchetterie total formulas
    for dech, positions in row_positions.items():
        if 'total_row' not in positions:
            continue
        
        total_row_num = positions['total_row']
        data_start = positions['data_start']
        data_end = positions['data_end']
        
        # Add formulas for category columns
        for col_idx, col_name in enumerate(all_columns, start=2):
            col_letter = get_column_letter(col_idx)
            ws.cell(row=total_row_num, column=col_idx).value = f'=SUM({col_letter}{data_start}:{col_letter}{data_end})'
        
        # TOTAL column formula
        col_idx = len(all_columns) + 2
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row_num, column=col_idx).value = f'=SUM(B{total_row_num}:{get_column_letter(col_idx-1)}{total_row_num})'
        
        # Sans massicot formula
        col_idx = len(all_columns) + 3
        # Find the actual column indices of MASSICOT and DEMANTELEMENT in all_columns
        massicot_col = None
        demantelement_col = None
        for idx, col_name in enumerate(all_columns, start=2):
            if col_name == 'MASSICOT':
                massicot_col = idx
            elif col_name == 'DEMANTELEMENT':
                demantelement_col = idx
        
        if massicot_col and demantelement_col:
            ws.cell(row=total_row_num, column=col_idx).value = (
                f'={get_column_letter(len(all_columns) + 2)}{total_row_num}'
                f'-{get_column_letter(massicot_col)}{total_row_num}'
                f'-{get_column_letter(demantelement_col)}{total_row_num}'
            )
    
    # Update grand total formulas
    for col_idx, col_name in enumerate(all_columns, start=2):
        col_letter = get_column_letter(col_idx)
        dech_total_rows = [positions['total_row'] for positions in row_positions.values() if 'total_row' in positions]
        if dech_total_rows:
            formula = '=SUM(' + ','.join([f'{col_letter}{r}' for r in dech_total_rows]) + ')'
            ws.cell(row=grand_total_row_num, column=col_idx).value = formula
    
    # TOTAL column in grand total
    col_idx = len(all_columns) + 2
    col_letter = get_column_letter(col_idx)
    dech_total_rows = [positions['total_row'] for positions in row_positions.values() if 'total_row' in positions]
    if dech_total_rows:
        formula = '=SUM(' + ','.join([f'{col_letter}{r}' for r in dech_total_rows]) + ')'
        ws.cell(row=grand_total_row_num, column=col_idx).value = formula
    
    # Sans massicot column in grand total
    col_idx = len(all_columns) + 3
    col_letter = get_column_letter(col_idx)
    # Find the actual column indices of MASSICOT and DEMANTELEMENT
    massicot_col = None
    demantelement_col = None
    for idx, col_name in enumerate(all_columns, start=2):
        if col_name == 'MASSICOT':
            massicot_col = idx
        elif col_name == 'DEMANTELEMENT':
            demantelement_col = idx
    
    if dech_total_rows and massicot_col and demantelement_col:
        formula = '=SUM(' + ','.join([f'{col_letter}{r}' for r in dech_total_rows]) + ')'
        ws.cell(row=grand_total_row_num, column=col_idx).value = formula
    
    # Add percentage formulas for each déchetterie
    # Store calculated values for gradient coloring using actual totals
    percentages_values = {}
    
    if percentages_data_rows:
        for dech, pct_row_num in percentages_data_rows.items():
            if dech not in row_positions:
                continue
            
            dech_total_row = row_positions[dech]['total_row']
            percentages_values[pct_row_num] = {}
            
            # Add percentage formulas for each category
            # Formula: (category total of this déchetterie) / (grand total of this category)
            for col_idx, col_name in enumerate(all_columns, start=2):
                col_letter = get_column_letter(col_idx)
                formula = f'=IF(${col_letter}${grand_total_row_num}=0,0,{col_letter}${dech_total_row}/${col_letter}${grand_total_row_num})'
                ws.cell(row=pct_row_num, column=col_idx).value = formula
                
                # Calculate value for gradient coloring using actual totals
                grand_total_val = grand_totals.get(col_name, 0)
                dech_total_val = dechetterie_totals[dech].get(col_name, 0)
                if grand_total_val != 0:
                    pct_value = dech_total_val / grand_total_val
                else:
                    pct_value = 0
                percentages_values[pct_row_num][col_idx] = pct_value
            
            # TOTAL column in percentages (should be 100% when summing across all déchetteries)
            col_idx = len(all_columns) + 2
            col_letter = get_column_letter(col_idx)
            formula = f'=IF(${col_letter}${grand_total_row_num}=0,0,{col_letter}${dech_total_row}/${col_letter}${grand_total_row_num})'
            ws.cell(row=pct_row_num, column=col_idx).value = formula
            
            grand_total_val = grand_totals['TOTAL']
            dech_total_val = sum(dechetterie_totals[dech].values()) - dechetterie_totals[dech].get('sans massicot et démantèlement', 0)
            if grand_total_val != 0:
                pct_value = dech_total_val / grand_total_val
            else:
                pct_value = 0
            percentages_values[pct_row_num][col_idx] = pct_value
            
            # Sans massicot et démantèlement column in percentages
            col_idx = len(all_columns) + 3
            col_letter = get_column_letter(col_idx)
            formula = f'=IF(${col_letter}${grand_total_row_num}=0,0,{col_letter}${dech_total_row}/${col_letter}${grand_total_row_num})'
            ws.cell(row=pct_row_num, column=col_idx).value = formula
            
            grand_total_val = grand_totals.get('sans massicot et démantèlement', 0)
            dech_total_val = dechetterie_totals[dech].get('sans massicot et démantèlement', 0)
            if grand_total_val != 0:
                pct_value = dech_total_val / grand_total_val
            else:
                pct_value = 0
            percentages_values[pct_row_num][col_idx] = pct_value
    
    # Format the worksheet for better readability
    print("   [OK] Mise en forme du tableau...")
    _format_worksheet(ws, all_columns, row_positions, grand_total_row_num, categories_reminder_row_num, 
                     percentages_title_row_num, percentages_header_row_num, percentages_data_rows, percentages_values)
    
    wb.save(output_file)
    
    # Final output
    print(f"\n[SUCCES] Fichier généré : {output_file}")
    print(f"\n{'='*70}")
    print(f"Grand Total du script: {grand_totals['TOTAL']/1000:,.2f} tonnes")
    print(f"{'='*70}")
    print("⚠️  À vérifier: Ce total DOIT correspondre au grand total du backend/frontend")
    print(f"{'='*70}\n")
    
    return output_file


def _format_worksheet(ws, all_columns, row_positions, grand_total_row_num, categories_reminder_row_num,
                     percentages_title_row_num=None, percentages_header_row_num=None, percentages_data_rows=None, percentages_values=None):
    """Apply formatting to the worksheet for better readability."""
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    
    def interpolate_color(percentage):
        """Interpolate color from white (0%) to blue (100%) based on percentage value."""
        if not isinstance(percentage, (int, float)) or percentage < 0:
            percentage = 0
        percentage = min(percentage, 1.0)  # Cap at 100%
        
        # Color interpolation: white (FFFFFF) to vivid blue (1F4E78)
        r_start, g_start, b_start = 255, 255, 255  # White
        r_end, g_end, b_end = 31, 78, 120  # Dark blue
        
        r = int(r_start + (r_end - r_start) * percentage)
        g = int(g_start + (g_end - g_start) * percentage)
        b = int(b_start + (b_end - b_start) * percentage)
        
        return f"{r:02X}{g:02X}{b:02X}"
    
    # Define colors
    color_title = "1F4E78"  # Dark blue
    color_header = "4472C4"  # Blue
    color_dechetterie = "D9E1F2"  # Light blue
    color_total = "FFC000"  # Gold
    color_grand_total = "FF6B6B"  # Red
    color_text = "FFFFFF"  # White
    color_dechets_ultimes = "FFD4E5"  # Pale pink (data rows)
    color_dechets_ultimes_total = "FF69B4"  # Vivid pink (total rows)
    color_massicot_demantelement = "D3D3D3"  # Light gray (data rows)
    color_massicot_demantelement_total = "808080"  # Dark gray (total rows)
    color_month_dechetterie_bg = "E8EEF7"  # Very light blue for month/déchetterie names
    
    # Find column indices for special columns
    dechets_ultimes_col = None
    massicot_col = None
    demantelement_col = None
    
    for idx, col_name in enumerate(all_columns, start=2):
        if col_name == 'DECHETS ULTIMES':
            dechets_ultimes_col = idx
        elif col_name == 'MASSICOT':
            massicot_col = idx
        elif col_name == 'DEMANTELEMENT':
            demantelement_col = idx
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Define fonts
    font_bold_white = Font(bold=True, color=color_text, size=11)
    font_bold = Font(bold=True, size=11)
    font_normal = Font(size=10)
    
    # Define alignment
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    # Format title row (row 1) - with better visibility
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14, color=color_text)
    title_cell.fill = PatternFill(start_color=color_title, end_color=color_title, fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 35
    
    # Merge title across columns
    ws.merge_cells(f'A1:{get_column_letter(len(all_columns) + 3)}1')
    
    # Add a thicker bottom border to the title for visual separation
    thick_border_bottom = Border(bottom=Side(style='medium', color='000000'))
    for col_idx in range(1, len(all_columns) + 4):
        ws.cell(row=1, column=col_idx).border = Border(
            left=Side(style='thin', color='000000') if col_idx == 1 else None,
            right=Side(style='thin', color='000000') if col_idx == len(all_columns) + 3 else None,
            top=Side(style='thin', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
    
    # Format percentages title row if it exists - with better visibility
    if percentages_title_row_num:
        pct_title_cell = ws.cell(row=percentages_title_row_num, column=1)
        pct_title_cell.font = Font(bold=True, size=14, color=color_text)
        pct_title_cell.fill = PatternFill(start_color=color_title, end_color=color_title, fill_type='solid')
        pct_title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[percentages_title_row_num].height = 35
        ws.merge_cells(f'A{percentages_title_row_num}:{get_column_letter(len(all_columns) + 3)}{percentages_title_row_num}')
        
        # Add a thicker bottom border to the percentages title for visual separation
        for col_idx in range(1, len(all_columns) + 4):
            ws.cell(row=percentages_title_row_num, column=col_idx).border = Border(
                left=Side(style='thin', color='000000') if col_idx == 1 else None,
                right=Side(style='thin', color='000000') if col_idx == len(all_columns) + 3 else None,
                top=Side(style='thin', color='000000'),
                bottom=Side(style='medium', color='000000')
            )
    
    # Format data
    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1)
        
        if not cell_a.value:
            ws.row_dimensions[row_idx].height = 15
            continue
        
        cell_a_str = str(cell_a.value).upper().strip() if cell_a.value else ""
        
        # Title row (already done)
        if row_idx == 1:
            continue
        
        # Check if it's a déchetterie header
        is_dechetterie_header = False
        for dech_name in [d.upper() for d in row_positions.keys()]:
            if cell_a_str == dech_name:
                is_dechetterie_header = True
                break
        
        # Check if it's a "Total" row for a déchetterie
        is_dechetterie_total = (
            cell_a_str == 'TOTAL' and
            any(row_idx == pos.get('total_row') for pos in row_positions.values())
        )
        
        # Check if it's the grand total row
        is_grand_total = (cell_a_str == 'TOTAL' and row_idx == grand_total_row_num)
        
        # Check if it's the categories reminder row
        is_categories_reminder = (cell_a_str == 'CATEGORIES' and row_idx == categories_reminder_row_num)
        
        # Check if it's the percentages header row
        is_percentages_header = (percentages_header_row_num and row_idx == percentages_header_row_num)
        
        # Check if it's a percentages data row
        is_percentages_data = (percentages_data_rows and any(row_idx == r for r in percentages_data_rows.values()))
        
        # Check if it's a month row
        month_list = ['JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN',
                      'JUILLET', 'AOUT', 'SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DECEMBRE']
        is_month_row = cell_a_str in month_list
        
        # Format rows
        if is_percentages_header:
            # Percentages header row - blue background with white text
            ws.row_dimensions[row_idx].height = 25
            for col_idx in range(1, len(all_columns) + 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type='solid')
                cell.font = font_bold_white
                cell.border = thin_border
                cell.alignment = align_center
        
        elif is_percentages_data:
            # Percentages data row - gradient color based on percentage value
            ws.row_dimensions[row_idx].height = 18
            cell_a.value = cell_a_str  # Ensure uppercase
            # Color the déchetterie name in column A
            cell_a.fill = PatternFill(start_color=color_month_dechetterie_bg, end_color=color_month_dechetterie_bg, fill_type='solid')
            cell_a.font = Font(bold=True, size=10, color="000000")
            cell_a.border = thin_border
            cell_a.alignment = align_left
            # Apply gradient colors to data columns (starting from col_idx=2)
            for col_idx in range(2, len(all_columns) + 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_normal
                cell.border = thin_border
                cell.alignment = align_right
                cell.number_format = '0.00%'
                
                # Apply gradient color based on percentage value from percentages_values
                if percentages_values and row_idx in percentages_values and col_idx in percentages_values[row_idx]:
                    pct_val = percentages_values[row_idx][col_idx]
                    if isinstance(pct_val, (int, float)) and pct_val > 0:
                        # Use gradient color based on percentage
                        gradient_color = interpolate_color(float(pct_val))
                        cell.fill = PatternFill(start_color=gradient_color, end_color=gradient_color, fill_type='solid')
                        # Adjust font color for readability
                        if float(pct_val) > 0.5:
                            cell.font = Font(bold=True, color="FFFFFF", size=10)  # White text for dark backgrounds
        
        elif is_categories_reminder:
            # Categories reminder row - blue background with white text for all columns
            ws.row_dimensions[row_idx].height = 40
            cell_a.value = 'CATÉGORIES'
            for col_idx in range(1, len(all_columns) + 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type='solid')
                cell.font = font_bold_white
                cell.border = thin_border
                cell.alignment = align_center
        
        elif is_grand_total:
            # Grand total row - red background
            ws.row_dimensions[row_idx].height = 22
            for col_idx in range(1, len(all_columns) + 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_bold_white
                cell.border = thin_border
                cell.alignment = align_right if col_idx > 1 else align_left
                if col_idx > 1:
                    cell.number_format = '#,##0.##'
                
                # Apply specific column colors (more vivid for totals)
                if col_idx == dechets_ultimes_col:
                    cell.fill = PatternFill(start_color=color_dechets_ultimes_total, end_color=color_dechets_ultimes_total, fill_type='solid')
                elif col_idx in [massicot_col, demantelement_col]:
                    cell.fill = PatternFill(start_color=color_massicot_demantelement_total, end_color=color_massicot_demantelement_total, fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color=color_grand_total, end_color=color_grand_total, fill_type='solid')
        
        elif is_dechetterie_total:
            # Déchetterie total row - gold background
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, len(all_columns) + 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_bold
                cell.border = thin_border
                cell.alignment = align_right if col_idx > 1 else align_left
                if col_idx > 1:
                    cell.number_format = '#,##0.##'
                
                # Apply specific column colors (more vivid for totals)
                if col_idx == dechets_ultimes_col:
                    cell.fill = PatternFill(start_color=color_dechets_ultimes_total, end_color=color_dechets_ultimes_total, fill_type='solid')
                elif col_idx in [massicot_col, demantelement_col]:
                    cell.fill = PatternFill(start_color=color_massicot_demantelement_total, end_color=color_massicot_demantelement_total, fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color=color_total, end_color=color_total, fill_type='solid')
        
        elif is_dechetterie_header:
            # Déchetterie header - blue background for ALL columns
            ws.row_dimensions[row_idx].height = 28
            cell_a.value = cell_a_str  # Ensure uppercase
            # Color the déchetterie name in column A
            cell_a.fill = PatternFill(start_color=color_month_dechetterie_bg, end_color=color_month_dechetterie_bg, fill_type='solid')
            cell_a.font = Font(bold=True, size=10, color="000000")
            cell_a.alignment = align_left
            for col_idx in range(1, len(all_columns) + 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type='solid')
                cell.font = font_bold_white
                cell.border = thin_border
                cell.alignment = align_center
        
        elif is_month_row:
            # Month data row - light blue background
            ws.row_dimensions[row_idx].height = 18
            cell_a.value = cell_a_str  # Ensure uppercase
            # Color the month name in column A
            cell_a.fill = PatternFill(start_color=color_month_dechetterie_bg, end_color=color_month_dechetterie_bg, fill_type='solid')
            cell_a.font = Font(bold=True, size=10)
            cell_a.border = thin_border
            cell_a.alignment = align_left
            for col_idx in range(1, len(all_columns) + 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_normal
                cell.border = thin_border
                cell.alignment = align_right if col_idx > 1 else align_left
                if col_idx > 1 and cell.value is not None and not isinstance(cell.value, str):
                    cell.number_format = '#,##0.##'
                
                # Only apply colors if value is not 0 or empty
                if cell.value is not None and cell.value != 0 and not isinstance(cell.value, str):
                    # Apply specific column colors
                    if col_idx == dechets_ultimes_col:
                        cell.fill = PatternFill(start_color=color_dechets_ultimes, end_color=color_dechets_ultimes, fill_type='solid')
                    elif col_idx in [massicot_col, demantelement_col]:
                        cell.fill = PatternFill(start_color=color_massicot_demantelement, end_color=color_massicot_demantelement, fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color=color_dechetterie, end_color=color_dechetterie, fill_type='solid')
        
        else:
            # Other rows
            ws.row_dimensions[row_idx].height = 15
            if cell_a.value:
                cell_a.value = str(cell_a.value).upper()
            for col_idx in range(1, len(all_columns) + 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = font_normal
                cell.alignment = align_right if col_idx > 1 else align_left
                if col_idx > 1 and cell.value is not None and not isinstance(cell.value, str):
                    cell.number_format = '#,##0.##'
                
                # Only apply colors if value is not 0 or empty
                if cell.value is not None and cell.value != 0 and not isinstance(cell.value, str):
                    # Apply specific column colors
                    if col_idx == dechets_ultimes_col:
                        cell.fill = PatternFill(start_color=color_dechets_ultimes, end_color=color_dechets_ultimes, fill_type='solid')
                    elif col_idx in [massicot_col, demantelement_col]:
                        cell.fill = PatternFill(start_color=color_massicot_demantelement, end_color=color_massicot_demantelement, fill_type='solid')
    
    # Set column widths
    ws.column_dimensions['A'].width = 20  # Déchetterie/Month column
    for col_idx in range(2, len(all_columns) + 2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 13  # Category columns
    
    # TOTAL column
    ws.column_dimensions[get_column_letter(len(all_columns) + 2)].width = 13
    # Sans massicot column
    ws.column_dimensions[get_column_letter(len(all_columns) + 3)].width = 18
    # Remaining columns
    for col_idx in range(len(all_columns) + 4, len(all_columns) + 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12


if __name__ == '__main__':
    project_root, input_dir, output_dir = _get_project_paths()
    
    # Create output directory if needed
    os.makedirs(output_dir, exist_ok=True)
    
    # Year for the dump
    year = 2025
    
    # Generate output filename
    output_filename = f"COLLECTES_RECYCLERIE_DUMP_{year}.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    # Run synthesis
    synthesize_dump(output_path, year=year)
