"""
Transform data from Excel file in 'input' folder (first sheet)
into summary format matching 'COLLECTES DECHETERIES 2025.xlsx'

This script processes data from ALL déchetteries (including special cases)
and can generate output for individual déchetteries or combined data.

The script automatically finds the first Excel file in the 'input' folder
and reads the first sheet, regardless of file or sheet names.
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION: Déchetterie Mapping
# ============================================================================
# Map from "Lieu collecte" values to standard déchetterie names
DECHETTERIE_MAPPING = {
    'Dech. La Pépiniere': 'Pépinière',
    'Dech. Sanssac': 'Sanssac',
    'Dech. Saint-Germain': 'St Germain',
    'Dech. Polignac': 'Polignac',
    # Special cases - keep as is or map to a standard name
    'Dech. Yssingeaux': 'Yssingeaux',
    'Dech. Bas-en-basset': 'Bas-en-basset',
    'Dech. Monistrol': 'Monistrol',
}

# Standard 4 déchetteries
STANDARD_DECHETTERIES = ['Pépinière', 'Sanssac', 'St Germain', 'Polignac']

# ============================================================================
# CONFIGURATION: Category Mapping Rules (same as before)
# ============================================================================
CATEGORY_MAPPING = {
    '4.CHINE': 'CHINE',
    '4.LIVRES': 'LIVRES',
    '4.MEUBLES': 'MEUBLES',
    '4.VAISSELLE': 'VAISSELLE',
    "4.BRICOLAGE ( EMMA'TEK)": 'ABJ',
    '4.JEUX/JOUETS': None,
    '4.PAM': None,
    '4.SPORTS-LOISIRS': None,
    '4.CADRES': 'CADRES',
    '4.MERCERIE': 'MERCERIE',
    '4.TEXTILES': 'TEXTILE',
    '4.PUERICULTURE': 'PUERICULTURE',
    '4.PAPETERIE': 'PAPETERIE',
    '4 .CD/DVD': 'CD/DVD/K7',
    '4.CHAUSSURES': None,  # Need to determine mapping
    '4.LABEL': None,  # Need to determine mapping
    '4.SACS': None,  # Need to determine mapping
}

def map_special_cases(categorie, sous_categorie, flux, orientation=None):
    """Handle special cases that need sub-category or flux information"""
    categorie_clean = str(categorie).upper().strip()
    sous_categorie_clean = str(sous_categorie).upper().strip()
    flux_clean = str(flux).upper().strip()
    # Note: orientation is already checked in map_category_to_collectes, so we don't check it here again
    
    if 'JEUX' in categorie_clean or 'JOUETS' in categorie_clean:
        if 'DEEE' in flux_clean or 'ELECTRIQUE' in sous_categorie_clean or 'ELECTRONIQUE' in sous_categorie_clean:
            return 'ELECTRO'
        else:
            return 'JOUETS'
    
    if 'PAM' in categorie_clean:
        if 'ECRAN' in sous_categorie_clean or 'LUMINAIRE' in sous_categorie_clean:
            return 'ELECTRO'
    
    if 'PAPIER' in flux_clean:
        return 'PAPETERIE'
    
    if 'ASL' in flux_clean:
        return 'ASL'
    
    if 'DEEE' in flux_clean:
        return 'ELECTRO'
    
    return None

def map_category_to_collectes(categorie, sous_categorie, flux, orientation=None):
    """Map categories from analyse format to COLLECTES format"""
    # Check orientation FIRST for MASSICOT and DEMANTELEMENT (these take priority)
    orientation_clean = str(orientation).upper().strip() if orientation else ''
    if 'MASICOT' in orientation_clean or 'MASSICOT' in orientation_clean:
        return 'MASSICOT'
    
    if 'DEMANTELLEMENT' in orientation_clean or 'DEMANTELEMENT' in orientation_clean:
        return 'DEMANTELEMENT'
    
    # Then check category mapping
    if categorie in CATEGORY_MAPPING:
        mapped = CATEGORY_MAPPING[categorie]
        if mapped is not None:
            return mapped
    
    # Finally check other special cases
    mapped = map_special_cases(categorie, sous_categorie, flux, orientation)
    if mapped is not None:
        return mapped
    
    return None

# ============================================================================
# Main Transformation Function
# ============================================================================

def transform_to_collectes(input_file, output_file, dechetterie_filter=None, combine_all=False):
    """
    Transform Excel data into COLLECTES summary format
    
    Args:
        input_file: Path to Excel file (will read first sheet)
        output_file: Path for output file
        dechetterie_filter: Specific déchetterie to process (None = all)
        combine_all: If True, combine all déchetteries into one output
    """
    
    print(f"\n{'='*70}")
    print(" " * 20 + "TRANSFORMATION EN COURS")
    print(f"{'='*70}")
    
    if not os.path.exists(input_file):
        print(f"\n❌ ERREUR : Fichier introuvable : {input_file}")
        return None
    
    # Read ALL sheets and combine them
    print(f"\n[LECTURE] Lecture du fichier : {os.path.basename(input_file)}")
    try:
        excel_file = pd.ExcelFile(input_file)
        all_sheets = excel_file.sheet_names
        print(f"   Feuilles detectees : {len(all_sheets)}")
        for sheet in all_sheets:
            print(f"     - {sheet}")
        
        # Read all sheets and combine them
        all_dataframes = []
        for sheet_name in all_sheets:
            try:
                sheet_df = pd.read_excel(input_file, sheet_name=sheet_name)
                # Check if this sheet has the required columns
                required_columns = ['Catégorie', 'Sous Catégorie', 'Flux', 'Poids', 'Date', 'Lieu collecte']
                if all(col in sheet_df.columns for col in required_columns):
                    all_dataframes.append(sheet_df)
                    print(f"   [OK] Feuille '{sheet_name}' : {len(sheet_df)} enregistrements avec colonnes requises")
                else:
                    print(f"   [SKIP] Feuille '{sheet_name}' : colonnes requises manquantes")
            except Exception as e:
                print(f"   [ERREUR] Impossible de lire la feuille '{sheet_name}' : {str(e)}")
                continue
        
        if not all_dataframes:
            print(f"\n[ERREUR] Aucune feuille valide trouvee dans le fichier")
            return None
        
        # Combine all dataframes
        df = pd.concat(all_dataframes, ignore_index=True)
        print(f"   [OK] Total combine : {len(df)} enregistrements depuis {len(all_dataframes)} feuille(s)")
    except Exception as e:
        print(f"\n[ERREUR] Impossible de lire le fichier Excel : {str(e)}")
        print(f"   Verifiez que le fichier n'est pas ouvert dans Excel.")
        return None
    
    # Check required columns
    required_columns = ['Catégorie', 'Sous Catégorie', 'Flux', 'Poids', 'Date', 'Lieu collecte']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        print(f"\n[ERREUR] Colonnes manquantes dans le fichier : {', '.join(missing_columns)}")
        print(f"   Le fichier Excel doit contenir ces colonnes.")
        return None
    
    # Filter by déchetterie if specified
    if dechetterie_filter:
        # Map standard name to "Lieu collecte" format
        lieu_mapping = {v: k for k, v in DECHETTERIE_MAPPING.items()}
        if dechetterie_filter in lieu_mapping:
            lieu_value = lieu_mapping[dechetterie_filter]
            df = df[df['Lieu collecte'] == lieu_value].copy()
            print(f"   Filtrage sur {dechetterie_filter} : {len(df)} enregistrements")
        else:
            # Try direct match
            df = df[df['Lieu collecte'].str.contains(dechetterie_filter, case=False, na=False)].copy()
            print(f"   Filtrage sur déchetterie contenant '{dechetterie_filter}' : {len(df)} enregistrements")
    
    if len(df) == 0:
        print("\n[ERREUR] Aucune donnee trouvee apres filtrage !")
        return None
    
    # Map déchetterie names
    df['Dechetterie'] = df['Lieu collecte'].map(DECHETTERIE_MAPPING)
    df['Dechetterie'] = df['Dechetterie'].fillna(df['Lieu collecte'])
    
    # Remove any NaN values
    df = df[df['Dechetterie'].notna()].copy()
    
    # Get unique déchetteries in data
    unique_dechetteries = [str(d) for d in df['Dechetterie'].unique() if pd.notna(d)]
    print(f"\n[DECHETTERIES] {len(unique_dechetteries)} dechetteries detectees :")
    for dech in unique_dechetteries:
        count = len(df[df['Dechetterie'] == dech])
        print(f"   - {dech} : {count} enregistrements")
    
    # Process dates
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    valid_dates = df['Date'].notna()
    df = df[valid_dates].copy()
    
    if len(df) == 0:
        print("\n[ERREUR] Aucune date valide trouvee !")
        return None
    
    min_date = df['Date'].min()
    max_date = df['Date'].max()
    date_range_str = f"DU {min_date.strftime('%d/%m/%Y')} au {max_date.strftime('%d/%m/%Y')}"
    print(f"\n[DATES] Periode des donnees : {date_range_str}")
    
    # Extract month names
    month_names_fr = {
        1: 'JANVIER', 2: 'FEVRIER', 3: 'MARS', 4: 'AVRIL',
        5: 'MAI', 6: 'JUIN', 7: 'JUILLET', 8: 'AOUT',
        9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DECEMBRE'
    }
    
    df['Month'] = df['Date'].dt.month
    df['MonthName'] = df['Month'].map(month_names_fr)
    df = df[df['MonthName'].notna()].copy()
    
    # Map categories
    print(f"\n[MAPPING] Mapping des categories...")
    df['MappedCategory'] = df.apply(
        lambda row: map_category_to_collectes(
            row['Catégorie'],
            row['Sous Catégorie'],
            row['Flux'],
            row.get('Orientation', None)
        ), axis=1
    )
    
    mapped_df = df[df['MappedCategory'].notna()].copy()
    unmapped_df = df[df['MappedCategory'].isna()].copy()
    
    print(f"   [OK] {len(mapped_df)} enregistrements mappes avec succes")
    if len(unmapped_df) > 0:
        print(f"   [ATTENTION] {len(unmapped_df)} enregistrements non mappes (peuvent etre ignores)")
        if len(unmapped_df) <= 20:  # Afficher seulement si peu nombreux
            print("\n   Categories non mappees :")
            unmapped_summary = unmapped_df.groupby(['Catégorie', 'Sous Catégorie', 'Flux']).size().reset_index(name='Count')
            print(unmapped_summary.to_string(index=False))
    
    # Keep original df for DECHETS ULTIMES (EVACUATION DECHETS is not mapped)
    original_df = df.copy()
    
    # Always create combined output with all déchetteries on one sheet
    print(f"\n[CREATION] Creation du fichier combine avec toutes les dechetteries...")
    return _create_combined_output_all_dechetteries(mapped_df, output_file, date_range_str, unique_dechetteries, original_df)

def _create_combined_output_all_dechetteries(mapped_df, output_file, date_range_str, unique_dechetteries, original_df=None):
    """Create output with all déchetteries on one sheet, with totals and statistics
    
    Args:
        mapped_df: DataFrame with mapped categories (for regular data)
        original_df: Original DataFrame before mapping (for DECHETS ULTIMES - EVACUATION DECHETS)
    """
    
    # Order déchetteries: standard first, then special cases
    standard_order = ['Pépinière', 'Sanssac', 'St Germain', 'Polignac']
    special_cases = [d for d in unique_dechetteries if d not in standard_order]
    ordered_dechetteries = [d for d in standard_order if d in unique_dechetteries] + sorted(special_cases)
    
    print(f"   Traitement de {len(ordered_dechetteries)} dechetteries...")
    
    # Standard columns (without déchetterie name)
    category_columns = ['MEUBLES', 'ELECTRO', 'DEMANTELEMENT', 'CHINE',
                       'VAISSELLE', 'JOUETS', 'PAPETERIE', 'LIVRES', 'MASSICOT',
                       'CADRES', 'ASL', 'PUERICULTURE', 'ABJ', 'CD/DVD/K7', 'PMCB',
                       'MERCERIE', 'TEXTILE']
    
    month_order = ['JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN',
                   'JUILLET', 'AOUT', 'SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DECEMBRE']
    
    all_sections = []
    grand_totals = {col: 0 for col in category_columns}
    grand_totals['TOTAL'] = 0
    grand_totals['sans massicot et démantèlement'] = 0
    
    # Track row numbers for each déchetterie's total row (for formulas)
    dechetterie_total_rows = {}
    dechetterie_start_rows = {}
    current_row = 4  # Start after header row
    
    # Process each déchetterie
    for dech in ordered_dechetteries:
        dech_df = mapped_df[mapped_df['Dechetterie'] == dech].copy()
        
        if len(dech_df) == 0:
            continue
        
        # Group by month and category
        summary = dech_df.groupby(['MonthName', 'MappedCategory'])['Poids'].sum().reset_index()
        
        # Pivot
        pivot_df = summary.pivot_table(
            index='MonthName',
            columns='MappedCategory',
            values='Poids',
            aggfunc='sum',
            fill_value=0
        ).reset_index()
        
        # Reorder months
        ordered_df = pd.DataFrame({'MonthName': month_order})
        result_df = ordered_df.merge(pivot_df, on='MonthName', how='left')
        result_df = result_df.fillna(0)
        
        # Add header row with déchetterie name
        dechetterie_header = dech.upper()
        header_row = [dechetterie_header] + category_columns + ['', '', '', 'DECHETS ULTIMES', '']
        all_sections.append(('header', header_row))
        dechetterie_start_rows[dech] = current_row
        current_row += 1  # Header row
        
        # Add data rows
        data_start_row = current_row
        for _, row in result_df.iterrows():
            data_row = [row['MonthName']]
            row_total = 0
            for col in category_columns:
                val = row[col] if col in row else 0
                data_row.append(val)
                if isinstance(val, (int, float)):
                    row_total += val
                    grand_totals[col] += val
            
            # TOTAL column
            data_row.append(row_total)
            grand_totals['TOTAL'] += row_total
            
            # sans massicot
            massicot = row.get('MASSICOT', 0) if 'MASSICOT' in row else 0
            demantelement = row.get('DEMANTELEMENT', 0) if 'DEMANTELEMENT' in row else 0
            sans_massicot = row_total - massicot - demantelement
            data_row.append(sans_massicot)
            grand_totals['sans massicot et démantèlement'] += sans_massicot
            
            # Empty columns (U, V, W) - V (index 21) reserved for DECHETS ULTIMES
            data_row.extend(['', '', ''])
            
            all_sections.append(('data', data_row))
            current_row += 1
        
        # Calculate DECHETS ULTIMES for this déchetterie by month
        # Use original_df because EVACUATION DECHETS is not mapped to a category
        if original_df is not None:
            dech_original = original_df[original_df['Dechetterie'] == dech].copy()
            ultimes_df = dech_original[
                (dech_original['Catégorie'] == 'EVACUATION DECHETS') & 
                (dech_original['Orientation'] == 'DECHETS ULTIMES')
            ].copy()
        else:
            # Fallback: try from mapped data (but EVACUATION DECHETS might not be mapped)
            ultimes_df = dech_df[
                (dech_df['Catégorie'] == 'EVACUATION DECHETS') & 
                (dech_df['Orientation'] == 'DECHETS ULTIMES')
            ].copy()
        
        if len(ultimes_df) > 0:
            # Ensure MonthName exists (it should from original_df processing)
            if 'MonthName' not in ultimes_df.columns:
                # Process dates if needed
                ultimes_df['Date'] = pd.to_datetime(ultimes_df['Date'], errors='coerce')
                month_names_fr = {
                    1: 'JANVIER', 2: 'FEVRIER', 3: 'MARS', 4: 'AVRIL',
                    5: 'MAI', 6: 'JUIN', 7: 'JUILLET', 8: 'AOUT',
                    9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DECEMBRE'
                }
                ultimes_df = ultimes_df[ultimes_df['Date'].notna()].copy()
                ultimes_df['Month'] = ultimes_df['Date'].dt.month
                ultimes_df['MonthName'] = ultimes_df['Month'].map(month_names_fr)
                ultimes_df = ultimes_df[ultimes_df['MonthName'].notna()].copy()
            
            # Group by month
            ultimes_summary = ultimes_df.groupby('MonthName')['Poids'].sum().reset_index()
            ultimes_dict = {}
            for _, ult_row in ultimes_summary.iterrows():
                month = ult_row['MonthName']
                ultimes_dict[month] = ult_row['Poids']
            all_sections.append(('ultimes', (dech, ultimes_dict)))
        
        # Add total row for this déchetterie (with formulas)
        total_row = ['Total']
        data_end_row = current_row - 1
        total_row_num = current_row
        dechetterie_total_rows[dech] = total_row_num
        
        # Add formulas for each category column
        for col_idx, col in enumerate(category_columns, start=2):  # Start at column B (index 2)
            col_letter = get_column_letter(col_idx)
            formula = f'=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})'
            total_row.append(formula)
        
        # TOTAL column formula
        col_start = get_column_letter(2)
        col_end = get_column_letter(18)
        total_row.append(f'=SUM({col_start}{total_row_num}:{col_end}{total_row_num})')
        
        # sans massicot formula
        total_row.append(f'=S{total_row_num}-J{total_row_num}-D{total_row_num}')
        # Columns U, V, W (V reserved for DECHETS ULTIMES total)
        total_row.extend(['', '', ''])
        total_row[21] = f'=SUM(V{data_start_row}:V{data_end_row})'
        all_sections.append(('total', total_row))
        current_row += 1
        
        # Empty row after each section
        all_sections.append(('empty', [''] * 23))
        current_row += 1
    
    # Create Excel file with all sections
    print(f"\n[ECRITURE] Ecriture du fichier Excel...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Title row
        title_text = f"COLLECTES DECHETERIES {date_range_str}"
        title_row = [[title_text] + [''] * 22]
        
        # Empty row with "sans massicot et démantèlement"
        empty_row = [[''] * 19 + ['sans massicot et démantèlement'] + [''] * 3]
        
        # Don't create a separate main header - the first déchetterie header will serve as the main header
        # Combine all rows and track row positions
        all_rows = title_row + empty_row
        row_positions = {}  # Track where each déchetterie's data starts/ends
        ultimes_data = {}  # Track DECHETS ULTIMES by déchetterie and month
        
        # Build rows and track positions
        current_excel_row = len(all_rows) + 1  # Excel row numbers (1-based)
        current_dechetterie = None
        
        for section_type, row_data in all_sections:
            if section_type == 'header':
                # Find which déchetterie this is
                dech_name = row_data[0] if row_data else ''
                # Find matching déchetterie
                for dech in ordered_dechetteries:
                    if dech.upper() == dech_name:
                        row_positions[dech] = {'header_row': current_excel_row, 'data_start': current_excel_row + 1}
                        current_dechetterie = dech
                        break
                all_rows.append(row_data[:23])
                current_excel_row += 1
            elif section_type == 'data':
                all_rows.append(row_data[:23])
                current_excel_row += 1
            elif section_type == 'total':
                # Find which déchetterie this total belongs to
                # It's the last one we processed
                for dech in reversed(ordered_dechetteries):
                    if dech in row_positions and 'total_row' not in row_positions[dech]:
                        row_positions[dech]['total_row'] = current_excel_row
                        row_positions[dech]['data_end'] = current_excel_row - 1
                        break
                all_rows.append(row_data[:23])
                current_excel_row += 1
            elif section_type == 'ultimes':
                # Store ultimes data
                dech, ultimes_dict = row_data
                ultimes_data[dech] = ultimes_dict
            else:  # empty
                all_rows.append(row_data[:23])
                current_excel_row += 1
        
        # Now update DECHETS ULTIMES in data rows
        # We need to update the actual list, not a copy
        month_order = ['JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN',
                       'JUILLET', 'AOUT', 'SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DECEMBRE']
        
        # Helper function to normalize names (remove accents for comparison)
        def normalize_name(name):
            """Normalize name by removing accents for comparison"""
            if not name:
                return ''
            name_upper = str(name).upper().strip()
            replacements = {
                'É': 'E', 'È': 'E', 'Ê': 'E', 'Ë': 'E',
                'À': 'A', 'Â': 'A', 'Ä': 'A',
                'Ù': 'U', 'Û': 'U', 'Ü': 'U',
                'Î': 'I', 'Ï': 'I',
                'Ô': 'O', 'Ö': 'O',
                'Ç': 'C'
            }
            for accented, unaccented in replacements.items():
                name_upper = name_upper.replace(accented, unaccented)
            return name_upper
        
        current_dechetterie = None
        start_idx = len(title_row + empty_row)
        
        for idx in range(start_idx, len(all_rows)):
            row = all_rows[idx]
            if not row or len(row) < 22:
                continue
                
            # Check if it's a header row (déchetterie name)
            if row[0] and isinstance(row[0], str):
                row0_normalized = normalize_name(row[0])
                # Check if it matches a déchetterie name
                for dech in ordered_dechetteries:
                    dech_normalized = normalize_name(dech)
                    if dech_normalized == row0_normalized:
                        current_dechetterie = dech
                        break
                continue
            
            # Check if it's a month row
            if row[0] and isinstance(row[0], str):
                row0_upper = str(row[0]).upper().strip()
                if row0_upper in [m.upper() for m in month_order]:
                    if current_dechetterie and current_dechetterie in ultimes_data:
                        month = row0_upper
                        if month in ultimes_data[current_dechetterie]:
                            all_rows[idx][21] = ultimes_data[current_dechetterie][month]  # Column 22 (index 21)
                continue
        
        # Add empty rows before statistics
        for _ in range(3):
            all_rows.append([''] * 23)
            current_excel_row += 1
        
        # Add grand total row - sum of all déchetterie totals
        grand_total_row = ['TOTAL']
        grand_total_row_num = len(all_rows) + 1  # Calculate based on actual row count
        for col_idx, col in enumerate(category_columns, start=2):
            col_letter = get_column_letter(col_idx)
            # Sum all déchetterie total rows
            formula_parts = []
            for dech in ordered_dechetteries:
                if dech in row_positions and 'total_row' in row_positions[dech]:
                    formula_parts.append(f'{col_letter}{row_positions[dech]["total_row"]}')
            if formula_parts:
                formula = '=' + '+'.join(formula_parts)
                grand_total_row.append(formula)
            else:
                grand_total_row.append(0)
        
        # TOTAL column formula
        col_start = get_column_letter(2)
        col_end = get_column_letter(18)
        grand_total_row.append(f'=SUM({col_start}{grand_total_row_num}:{col_end}{grand_total_row_num})')
        grand_total_row.append('')  # sans massicot (can be calculated if needed)
        # Columns U, V, W (V reserved for DECHETS ULTIMES total)
        ultimes_formula_parts = []
        for dech in ordered_dechetteries:
            if dech in row_positions and 'total_row' in row_positions[dech]:
                ultimes_formula_parts.append(f'V{row_positions[dech]["total_row"]}')
        ultimes_total_formula = '=' + '+'.join(ultimes_formula_parts) if ultimes_formula_parts else '0'
        grand_total_row.extend(['', '', ''])
        grand_total_row[21] = ultimes_total_formula
        all_rows.append(grand_total_row[:23])
        
        # Add empty row
        all_rows.append([''] * 23)
        
        # Add percentage rows for each déchetterie (PEP, POL, STG, SAN, etc.)
        dech_abbrev = {
            'Pépinière': 'PEP',
            'Polignac': 'POL',
            'St Germain': 'STG',
            'Sanssac': 'SAN',
            'Yssingeaux': 'YSS',
            'Bas-en-basset': 'BAS',
            'Monistrol': 'MON'
        }
        
        for dech in ordered_dechetteries:
            if dech in row_positions and 'total_row' in row_positions[dech] and dech in dech_abbrev:
                abbrev = dech_abbrev[dech]
                total_row_num = row_positions[dech]['total_row']
                percent_row = [abbrev]
                for col_idx, col in enumerate(category_columns, start=2):
                    col_letter = get_column_letter(col_idx)
                    # Formula should return decimal (0.46), Excel will display as 46% with percentage format
                    # Don't use + prefix to avoid any potential issues
                    formula = f'={col_letter}{total_row_num}/{col_letter}{grand_total_row_num}'
                    percent_row.append(formula)
                percent_row.extend([''] * 5)  # Fill remaining columns
                all_rows.append(percent_row[:23])
        
        # Add empty rows
        for _ in range(2):
            all_rows.append([''] * 23)
        
        # Add summary rows
        # TOTAL COLLECTES X DECHETERIES
        summary_row1 = [''] * 4 + ['TOTAL COLLECTES'] + [''] + [str(len(ordered_dechetteries))] + [''] * 15
        all_rows.append(summary_row1[:23])
        
        # Empty row
        all_rows.append([''] * 23)
        
        # DECHETS ULTIMES rows (placeholder - user can fill manually)
        # These are typically filled manually based on specific data
        ultimes_row1 = [''] * 4 + ['DECHETS ULTIMES T1'] + [''] + [''] + [''] * 15
        ultimes_row2 = [''] * 4 + ['DECHETS ULTIMES T2'] + [''] + [''] + [''] * 15
        all_rows.append(ultimes_row1[:23])
        all_rows.append(ultimes_row2[:23])
        
        # Empty rows
        for _ in range(2):
            all_rows.append([''] * 23)
        
        # MASSICOT and DEMANTELEMENT summary rows
        massicot_col = get_column_letter(10)  # Column J
        demantelement_col = get_column_letter(4)  # Column D
        massicot_row = [''] * 4 + ['MASSICOT'] + [''] + [f'={massicot_col}{grand_total_row_num}'] + [''] * 15
        demantelement_row = [''] * 4 + ['DEMANTELEMENT'] + [''] + [f'={demantelement_col}{grand_total_row_num}'] + [''] * 15
        all_rows.append(massicot_row[:23])
        all_rows.append(demantelement_row[:23])
        
        # Write to Excel
        final_df = pd.DataFrame(all_rows)
        final_df.to_excel(writer, sheet_name='Feuil1', index=False, header=False)
        final_df.to_excel(writer, sheet_name='CALCUL POIDS', index=False, header=False)
    
    # Now update formulas in the Excel file with correct row numbers
    print("   [OK] Mise a jour des formules...")
    wb = load_workbook(output_file)
    sheet_names = [name for name in ['Feuil1', 'CALCUL POIDS'] if name in wb.sheetnames]

    # Helper function to normalize names (remove accents for comparison)
    def normalize_name(name):
        """Normalize name by removing accents for comparison"""
        if not name:
            return ''
        name_upper = str(name).upper().strip()
        replacements = {
            'É': 'E', 'È': 'E', 'Ê': 'E', 'Ë': 'E',
            'À': 'A', 'Â': 'A', 'Ä': 'A',
            'Ù': 'U', 'Û': 'U', 'Ü': 'U',
            'Î': 'I', 'Ï': 'I',
            'Ô': 'O', 'Ö': 'O',
            'Ç': 'C'
        }
        for accented, unaccented in replacements.items():
            name_upper = name_upper.replace(accented, unaccented)
        return name_upper

    # First, collect all déchetterie total row numbers
    dechetterie_total_rows = []
    for dech in ordered_dechetteries:
        if dech in row_positions and 'total_row' in row_positions[dech]:
            dechetterie_total_rows.append(row_positions[dech]['total_row'])

    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        # Update DECHETS ULTIMES in Excel file (after writing)
        if ultimes_data:
            print(f"   [OK] Mise a jour des donnees DECHETS ULTIMES ({sheet_name})...")
            month_order = ['JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN',
                           'JUILLET', 'AOUT', 'SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DECEMBRE']

            current_dechetterie = None

            for row_idx in range(1, ws.max_row + 1):
                cell_a = ws.cell(row=row_idx, column=1)
                if not cell_a.value:
                    continue

                cell_a_str = str(cell_a.value)
                cell_a_normalized = normalize_name(cell_a_str)

                # Check if it's a déchetterie header
                for dech in ordered_dechetteries:
                    dech_normalized = normalize_name(dech)
                    if dech_normalized == cell_a_normalized:
                        current_dechetterie = dech
                        break

                # Check if it's a month row
                if cell_a_str.upper() in [m.upper() for m in month_order]:
                    if current_dechetterie and current_dechetterie in ultimes_data:
                        month = cell_a_str.upper()
                        if month in ultimes_data[current_dechetterie]:
                            cell_v = ws.cell(row=row_idx, column=22)
                            cell_v.value = ultimes_data[current_dechetterie][month]

        # Find grand total row number
        # The grand total row should have formulas that sum multiple déchetterie total rows
        # Look for a TOTAL row that has formulas like =B17+B32+B47+B62 (sum of multiple total rows)
        grand_total_row_num = None

        # Now find the grand total row - it should have formulas that reference multiple déchetterie totals
        for row_idx in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1)
            if cell_a.value and isinstance(cell_a.value, str) and cell_a.value.upper() == 'TOTAL':
                # Check if this row's formulas reference multiple déchetterie totals
                cell_b = ws.cell(row=row_idx, column=2)
                if cell_b.value and isinstance(cell_b.value, str) and cell_b.value.startswith('='):
                    # Count how many déchetterie total rows are referenced in this formula
                    formula = cell_b.value
                    referenced_count = sum(1 for total_row in dechetterie_total_rows if f'{get_column_letter(2)}{total_row}' in formula)
                    # If it references multiple déchetterie totals, it's likely the grand total
                    if referenced_count >= 2:
                        grand_total_row_num = row_idx
                        break

        # Fallback: find the TOTAL row that comes after all déchetterie sections
        if grand_total_row_num is None:
            max_dechetterie_total = max(dechetterie_total_rows) if dechetterie_total_rows else 0
            for row_idx in range(max_dechetterie_total + 5, ws.max_row + 1):
                cell_a = ws.cell(row=row_idx, column=1)
                if cell_a.value and isinstance(cell_a.value, str) and cell_a.value.upper() == 'TOTAL':
                    cell_b = ws.cell(row=row_idx, column=2)
                    if cell_b.value and isinstance(cell_b.value, str) and ('+' in cell_b.value or 'SUM' in cell_b.value):
                        grand_total_row_num = row_idx
                        break

        # Update total row formulas for each déchetterie
        for dech in ordered_dechetteries:
            if dech in row_positions:
                pos = row_positions[dech]
                if 'total_row' in pos and 'data_start' in pos and 'data_end' in pos:
                    total_row = pos['total_row']
                    data_start = pos['data_start']
                    data_end = pos['data_end']

                    # Update category column formulas
                    for col_idx, col in enumerate(category_columns, start=2):
                        col_letter = get_column_letter(col_idx)
                        cell = ws.cell(row=total_row, column=col_idx)
                        cell.value = f'=SUM({col_letter}{data_start}:{col_letter}{data_end})'

                    # Update TOTAL column formula
                    col_start = get_column_letter(2)
                    col_end = get_column_letter(18)
                    cell = ws.cell(row=total_row, column=19)
                    cell.value = f'=SUM({col_start}{total_row}:{col_end}{total_row})'

                    # Update sans massicot formula
                    cell = ws.cell(row=total_row, column=20)
                    cell.value = f'=S{total_row}-J{total_row}-D{total_row}'

                    # Update DECHETS ULTIMES total (column V)
                    cell = ws.cell(row=total_row, column=22)
                    cell.value = f'=SUM(V{data_start}:V{data_end})'

        # Update percentage formulas if grand total row was found
        if grand_total_row_num:
            # Update grand total for DECHETS ULTIMES (column V)
            ultimes_formula_parts = [f'V{row_num}' for row_num in dechetterie_total_rows]
            ultimes_total_formula = '=' + '+'.join(ultimes_formula_parts) if ultimes_formula_parts else '0'
            cell = ws.cell(row=grand_total_row_num, column=22)
            cell.value = ultimes_total_formula

            # Also write the total in the "DECHETS ULTIMES T2" summary row (column G)
            for row_idx in range(1, ws.max_row + 1):
                cell_label = ws.cell(row=row_idx, column=5)
                if cell_label.value and isinstance(cell_label.value, str) and cell_label.value.strip().upper() == 'DECHETS ULTIMES T2':
                    ws.cell(row=row_idx, column=7).value = ultimes_total_formula
                    break

            dech_abbrev = {
                'Pépinière': 'PEP',
                'Polignac': 'POL',
                'St Germain': 'STG',
                'Sanssac': 'SAN',
                'Yssingeaux': 'YSS',
                'Bas-en-basset': 'BAS',
                'Monistrol': 'MON'
            }

            for row_idx in range(1, ws.max_row + 1):
                cell_a = ws.cell(row=row_idx, column=1)
                if cell_a.value and isinstance(cell_a.value, str):
                    abbrev = cell_a.value.upper()
                    # Find matching déchetterie
                    for dech in ordered_dechetteries:
                        if dech in dech_abbrev and dech_abbrev[dech] == abbrev:
                            if dech in row_positions and 'total_row' in row_positions[dech]:
                                total_row_num = row_positions[dech]['total_row']
                                # Update percentage formulas
                                # Note: Excel percentage format automatically multiplies by 100
                                # So the formula should return a decimal (e.g., 0.46 for 46%)
                                for col_idx, col in enumerate(category_columns, start=2):
                                    col_letter = get_column_letter(col_idx)
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    # Formula should return decimal (0.46), Excel will display as 46% with percentage format
                                    cell.value = f'={col_letter}{total_row_num}/{col_letter}{grand_total_row_num}'
                                    cell.number_format = '0.00%'
                            break

    wb.save(output_file)
    wb.close()
    
    # Apply formatting
    print("   [OK] Application du formatage (couleurs, bordures)...")
    _apply_formatting_to_combined_file(output_file, len(ordered_dechetteries), len(month_order), row_positions, ordered_dechetteries)
    
    print(f"\n   [OK] Fichier cree avec succes !")
    return mapped_df

def _apply_formatting_to_combined_file(output_file, num_dechetteries, num_months, row_positions, ordered_dechetteries):
    """Apply formatting to combined file with multiple déchetteries"""
    wb = load_workbook(output_file)
    
    # Define styles
    medium_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Different colors for each déchetterie
    dechetterie_colors = {
        'Pépinière': 'E2EFDA',  # Light green
        'Sanssac': 'FFF2CC',     # Light yellow
        'St Germain': 'DEEBF7',  # Light blue
        'Polignac': 'F4B084',    # Light orange
        'Yssingeaux': 'D9D2E9',  # Light purple
        'Bas-en-basset': 'FCE4D6', # Light peach
        'Monistrol': 'C5E0B4',   # Light green-yellow
    }
    
    # Gray color for DÉMANTÈLEMENT and MASSICOT columns
    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    # Standard data fill (will be overridden by déchetterie colors)
    data_fill_default = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    data_font = Font(name='Calibri', size=11, bold=False)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    month_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    month_font = Font(name='Calibri', size=11, bold=False)
    month_alignment = Alignment(horizontal='center', vertical='center')
    
    total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    total_font = Font(name='Calibri', size=12, bold=True)
    total_alignment = Alignment(horizontal='center', vertical='center')
    
    # Category columns order (to find DEMANTELEMENT and MASSICOT column indices)
    category_columns = ['MEUBLES', 'ELECTRO', 'DEMANTELEMENT', 'CHINE',
                       'VAISSELLE', 'JOUETS', 'PAPETERIE', 'LIVRES', 'MASSICOT',
                       'CADRES', 'ASL', 'PUERICULTURE', 'ABJ', 'CD/DVD/K7', 'PMCB',
                       'MERCERIE', 'TEXTILE']
    
    # Find column indices for DEMANTELEMENT (index 2, column D) and MASSICOT (index 8, column J)
    demantelement_col_idx = 4  # Column D (DEMANTELEMENT is 3rd category, so column 2+2=4)
    massicot_col_idx = 10      # Column J (MASSICOT is 9th category, so column 2+8=10)
    
    for sheet_name in ['Feuil1', 'CALCUL POIDS']:
        ws = wb[sheet_name]
        
        # Set row heights
        ws.row_dimensions[1].height = 26.25
        ws.row_dimensions[2].height = 15.75
        ws.row_dimensions[3].height = 15.75
        
        # Format Row 3 (main header)
        for col in range(1, 19):
            cell = ws.cell(row=3, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = medium_border
        
        cell = ws.cell(row=3, column=22)
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Format all data rows
        current_dechetterie = None
        current_dechetterie_color = None
        
        # Helper function to normalize déchetterie names
        def normalize_dechetterie_name(name):
            """Normalize déchetterie name for matching"""
            if not name:
                return ''
            name_upper = str(name).upper().strip()
            replacements = {
                'É': 'E', 'È': 'E', 'Ê': 'E', 'Ë': 'E',
                'À': 'A', 'Â': 'A', 'Ä': 'A',
                'Ù': 'U', 'Û': 'U', 'Ü': 'U',
                'Î': 'I', 'Ï': 'I',
                'Ô': 'O', 'Ö': 'O',
                'Ç': 'C'
            }
            for accented, unaccented in replacements.items():
                name_upper = name_upper.replace(accented, unaccented)
            return name_upper
        
        for row_idx in range(4, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1)
            val_a = cell_a.value
            
            if val_a is None or val_a == '':
                continue
            
            val_str = str(val_a).upper()
            
            # Check if it's a déchetterie header
            val_normalized = normalize_dechetterie_name(val_a)
            found_dechetterie = None
            for dech in ordered_dechetteries:
                dech_normalized = normalize_dechetterie_name(dech)
                if dech_normalized == val_normalized:
                    found_dechetterie = dech
                    break
            
            if found_dechetterie:
                # Header row - format all columns
                current_dechetterie = found_dechetterie
                current_dechetterie_color = dechetterie_colors.get(found_dechetterie, 'FFFF00')  # Default yellow
                data_fill = PatternFill(start_color=current_dechetterie_color, end_color=current_dechetterie_color, fill_type='solid')
                
                for col in range(1, 19):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = medium_border
                ws.row_dimensions[row_idx].height = 15.75
                continue
            
            # Check if it's a month name
            if val_str in ['JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN', 'JUILLET', 'AOUT', 'SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DECEMBRE']:
                # Month row
                cell_a.fill = month_fill
                cell_a.font = month_font
                cell_a.alignment = month_alignment
                cell_a.border = medium_border
                ws.row_dimensions[row_idx].height = 16.5
                
                # Use current déchetterie color, or default
                if current_dechetterie_color:
                    data_fill = PatternFill(start_color=current_dechetterie_color, end_color=current_dechetterie_color, fill_type='solid')
                else:
                    data_fill = data_fill_default
                
                # Data columns (2-18, but column 1 is month name)
                for col in range(2, 19):
                    cell = ws.cell(row=row_idx, column=col)
                    val = cell.value
                    if val is not None and val != '':
                        # Gray for DEMANTELEMENT (col 4) and MASSICOT (col 10)
                        if col == demantelement_col_idx or col == massicot_col_idx:
                            cell.fill = gray_fill
                        else:
                            cell.fill = data_fill
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = medium_border
                
                # TOTAL column
                cell = ws.cell(row=row_idx, column=19)
                if cell.value is not None and cell.value != '':
                    col_start = get_column_letter(2)
                    col_end = get_column_letter(18)
                    cell.value = f'=SUM({col_start}{row_idx}:{col_end}{row_idx})'
                    cell.fill = total_fill
                    cell.font = total_font
                    cell.alignment = total_alignment
                    cell.border = medium_border
                
                # sans massicot
                cell = ws.cell(row=row_idx, column=20)
                if cell.value is not None and cell.value != '':
                    cell.value = f'=S{row_idx}-J{row_idx}-D{row_idx}'
                    cell.font = total_font
                    cell.alignment = total_alignment
                
                continue
            
            # Check if it's a total row
            if 'TOTAL' in val_str and val_str != 'TOTAL GENERAL':
                # Total row for a déchetterie - bold, larger font
                for col in range(1, 20):
                    cell = ws.cell(row=row_idx, column=col)
                    if col == 1:
                        cell.font = total_font
                        cell.alignment = total_alignment
                    elif col < 19:
                        val = cell.value
                        if val is not None and val != '':
                            # Gray for DEMANTELEMENT (col 4) and MASSICOT (col 10)
                            if col == demantelement_col_idx or col == massicot_col_idx:
                                cell.fill = gray_fill
                            else:
                                cell.fill = total_fill
                            cell.font = total_font
                            cell.alignment = total_alignment
                            cell.border = medium_border
                    elif col == 19:
                        cell.font = total_font
                        cell.alignment = total_alignment
                        cell.border = medium_border
                ws.row_dimensions[row_idx].height = 16.5
                continue
            
            # Check if it's a percentage row (PEP, POL, STG, SAN, etc.)
            if val_str in ['PEP', 'POL', 'STG', 'SAN', 'YSS', 'BAS', 'MON']:
                # Percentage row - format as percentage
                for col in range(2, 19):
                    cell = ws.cell(row=row_idx, column=col)
                    if cell.value:
                        # Apply percentage format - Excel will multiply by 100 automatically
                        # So if formula returns 0.46, it will display as 46%
                        cell.number_format = '0.00%'
                        # Gray for DEMANTELEMENT (col 4) and MASSICOT (col 10)
                        if col == demantelement_col_idx or col == massicot_col_idx:
                            cell.fill = gray_fill
                        cell.font = data_font
                        cell.alignment = data_alignment
                ws.row_dimensions[row_idx].height = 16.5
                continue
            
            # Check if it's grand total row
            if val_str == 'TOTAL' and row_idx > 60:  # Grand total is usually later
                for col in range(1, 20):
                    cell = ws.cell(row=row_idx, column=col)
                    if col == 1:
                        cell.font = total_font
                        cell.alignment = total_alignment
                    elif col < 19:
                        val = cell.value
                        if val is not None and val != '':
                            # Gray for DEMANTELEMENT (col 4) and MASSICOT (col 10)
                            if col == demantelement_col_idx or col == massicot_col_idx:
                                cell.fill = gray_fill
                            else:
                                cell.fill = total_fill
                            cell.font = total_font
                            cell.alignment = total_alignment
                            cell.border = medium_border
                    elif col == 19:
                        cell.font = total_font
                        cell.alignment = total_alignment
                        cell.border = medium_border
                ws.row_dimensions[row_idx].height = 16.5
                continue
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        for col in range(2, 19):
            ws.column_dimensions[get_column_letter(col)].width = 10
        ws.column_dimensions['S'].width = 12
        ws.column_dimensions['T'].width = 12
    
    wb.save(output_file)
    wb.close()

if __name__ == "__main__":
    import sys
    
    print("="*70)
    print(" " * 15 + "OUTIL DE TRANSFORMATION DECHETTERIES")
    print("="*70)
    print("\nCe script transforme les données détaillées en format COLLECTES")
    print("pour présentation client.\n")
    
    # Aide si demandée
    if len(sys.argv) > 1 and sys.argv[1] in ['--help', '-h', '--aide', '/?']:
        print("UTILISATION:")
        print("  python transform_collectes.py")
        print("\n  Ou avec un nom de fichier de sortie personnalisé:")
        print("  python transform_collectes.py mon_fichier.xlsx")
        print("\nCE QUE FAIT LE SCRIPT:")
        print("  - Lit automatiquement le premier fichier Excel trouve dans le dossier 'input'")
        print("  - Utilise toujours la premiere feuille du fichier Excel")
        print("  - Combine TOUTES les déchetteries sur une seule feuille")
        print("  - Crée un fichier Excel formaté dans le dossier 'output'")
        print("  - Inclut les totaux et statistiques pour chaque déchetterie")
        print("\nFICHIERS NÉCESSAIRES:")
        print("  - Un fichier Excel (.xlsx ou .xls) dans le dossier 'input'")
        print("  - Le fichier doit contenir les colonnes requises (Catégorie, Flux, Poids, etc.)")
        print("\nFICHIER GÉNÉRÉ:")
        print("  - output/COLLECTES DECHETERIES T2 2025.xlsx")
        print("="*70)
        sys.exit(0)
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    input_dir = os.path.join(project_root, 'input')
    output_dir = os.path.join(project_root, 'output')
    
    # Créer les dossiers si nécessaire
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    
    # Chercher automatiquement le premier fichier Excel dans le dossier input
    input_file = None
    excel_extensions = ['.xlsx', '.xls']
    
    if os.path.exists(input_dir):
        for file in os.listdir(input_dir):
            file_path = os.path.join(input_dir, file)
            if os.path.isfile(file_path):
                _, ext = os.path.splitext(file)
                if ext.lower() in excel_extensions:
                    input_file = file_path
                    break
    
    # Vérifier que le fichier d'entrée existe
    if input_file is None or not os.path.exists(input_file):
        print(f"\n{'='*70}")
        print("ERREUR : Aucun fichier Excel trouve dans le dossier 'input' !")
        print(f"{'='*70}")
        print(f"\nLe dossier 'input' doit contenir un fichier Excel (.xlsx ou .xls)")
        print(f"  Dossier recherche : {input_dir}")
        print(f"\nVÉRIFICATIONS :")
        print(f"  1. Y a-t-il un fichier Excel dans le dossier 'input' ?")
        print(f"  2. Le fichier n'est-il pas ouvert dans Excel ?")
        print(f"  3. Le fichier a-t-il une extension .xlsx ou .xls ?")
        print(f"\nPour obtenir de l'aide, tapez :")
        print(f"  python transform_collectes.py --help")
        print(f"{'='*70}\n")
        sys.exit(1)
    
    # Nom du fichier de sortie
    if len(sys.argv) > 1:
        output_file = sys.argv[1]
        if not os.path.isabs(output_file):
            output_file = os.path.join(output_dir, output_file)
    else:
        output_file = os.path.join(output_dir, "COLLECTES DECHETERIES T2 2025.xlsx")
    
    print(f"Fichier d'entrée  : {input_file}")
    print(f"Fichier de sortie : {output_file}")
    print(f"\nTraitement en cours...\n")
    
    try:
        result = transform_to_collectes(input_file, output_file, None, combine_all=True)
        
        if result is not None:
            print(f"\n{'='*70}")
            print(" " * 25 + "SUCCES !")
            print(f"{'='*70}")
            print(f"\nLe fichier a ete cree avec succes :")
            print(f"  {output_file}")
            print(f"\nCe fichier contient :")
            print(f"  - Toutes les dechetteries sur une seule feuille")
            print(f"  - Les totaux pour chaque dechetterie")
            print(f"  - Les statistiques et pourcentages")
            print(f"  - Le formatage professionnel (couleurs, bordures)")
            print(f"\nVous pouvez maintenant ouvrir ce fichier dans Excel.")
            print(f"{'='*70}\n")
        else:
            print(f"\n{'='*70}")
            print(" " * 25 + "ERREUR")
            print(f"{'='*70}")
            print(f"\nLa transformation a echoue.")
            print(f"Verifiez les messages d'erreur ci-dessus.")
            print(f"{'='*70}\n")
            sys.exit(1)
            
    except FileNotFoundError as e:
        print(f"\n{'='*70}")
        print("ERREUR : Fichier introuvable")
        print(f"{'='*70}")
        print(f"\n{str(e)}")
        print(f"\nVerifiez que tous les fichiers necessaires sont dans le bon dossier.")
        print(f"{'='*70}\n")
        sys.exit(1)
    except Exception as e:
        print(f"\n{'='*70}")
        print("ERREUR : Une erreur inattendue s'est produite")
        print(f"{'='*70}")
        print(f"\nType d'erreur : {type(e).__name__}")
        print(f"Message : {str(e)}")
        print(f"\nSi le problème persiste, contactez le support technique.")
        print(f"{'='*70}\n")
        sys.exit(1)
