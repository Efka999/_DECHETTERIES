"""
Service pour extraire les statistiques du fichier Excel de sortie
"""

import pandas as pd
import numpy as np
import re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook


def _extract_date_range(ws):
    """Extract date range from the title row, if present."""
    if not ws or ws.max_row < 1:
        return None, None
    title_value = ws.cell(row=1, column=1).value
    if not title_value:
        return None, None
    title_text = str(title_value)
    match = re.search(r'DU\s+(\d{2}/\d{2}/\d{4})\s+au\s+(\d{2}/\d{2}/\d{4})', title_text, re.IGNORECASE)
    if not match:
        return None, None
    start_str, end_str = match.groups()
    try:
        start_date = datetime.strptime(start_str, '%d/%m/%Y')
        end_date = datetime.strptime(end_str, '%d/%m/%Y')
        return start_date, end_date
    except ValueError:
        return None, None


def parse_output_excel(file_path):
    """
    Parse le fichier Excel de sortie et extrait les statistiques
    
    Args:
        file_path: Chemin vers le fichier Excel de sortie
        
    Returns:
        dict: Statistiques structurées
    """
    try:
        # Lire SEULEMENT la première feuille (Feuil1) - la deuxième feuille est une copie
        # Utiliser openpyxl directement pour un meilleur contrôle
        wb = load_workbook(file_path, data_only=True)
        
        # S'assurer qu'on lit seulement la première feuille
        if 'Feuil1' in wb.sheetnames:
            ws = wb['Feuil1']
        else:
            # Fallback sur la première feuille disponible
            ws = wb.active
        
        category_columns = ['MEUBLES', 'ELECTRO', 'DEMANTELEMENT', 'CHINE',
                          'VAISSELLE', 'JOUETS', 'PAPETERIE', 'LIVRES', 'MASSICOT',
                          'CADRES', 'ASL', 'PUERICULTURE', 'ABJ', 'CD/DVD/K7', 'PMCB',
                          'MERCERIE', 'TEXTILE']
        final_fluxes = ['MASSICOT', 'DEMANTELEMENT', 'DECHETS ULTIMES']

        # Extraire la période depuis le titre (si disponible)
        date_start, date_end = _extract_date_range(ws)
        dataset_year = None
        date_range_label = None
        if date_start and date_end:
            date_range_label = f"DU {date_start.strftime('%d/%m/%Y')} au {date_end.strftime('%d/%m/%Y')}"
            if date_start.year == date_end.year:
                dataset_year = str(date_start.year)
            else:
                dataset_year = f"{date_start.year}-{date_end.year}"
        
        # Trouver les indices de colonnes pour chaque catégorie en cherchant dans les headers
        # On va chercher la ligne header qui contient les noms de colonnes
        column_indices = {}  # {category_name: column_index}
        header_row_found = False
        
        ultimes_col_idx = None
        # Chercher la ligne header (contient les noms de catégories)
        for row_idx in range(1, min(10, ws.max_row + 1)):  # Chercher dans les 10 premières lignes
            row = [cell.value for cell in ws[row_idx]]
            row_upper = [str(cell).upper().strip() if cell else '' for cell in row]
            
            # Vérifier si cette ligne contient des noms de catégories
            found_categories = [col for col in category_columns if col in row_upper]
            if len(found_categories) >= 5:  # Si on trouve au moins 5 catégories, c'est probablement le header
                # Mapper les colonnes
                for col_idx, cell_value in enumerate(row, start=1):
                    if cell_value:
                        cell_upper = str(cell_value).upper().strip()
                        # Chercher la correspondance dans category_columns
                        for cat_col in category_columns:
                            if cat_col == cell_upper or cat_col in cell_upper:
                                column_indices[cat_col] = col_idx
                                break
                        if 'DECHETS ULTIMES' in cell_upper:
                            ultimes_col_idx = col_idx
                header_row_found = True
                break
        
        # Si on n'a pas trouvé le header, utiliser l'ordre par défaut (colonne 2 = MEUBLES, etc.)
        if not header_row_found or len(column_indices) < len(category_columns):
            # Ordre par défaut : colonne 1 = déchetterie/mois, colonne 2+ = catégories
            for idx, cat_col in enumerate(category_columns, start=2):
                if cat_col not in column_indices:
                    column_indices[cat_col] = idx
        if ultimes_col_idx is None:
            ultimes_col_idx = 22
        
        # Extraire les données par déchetterie
        dechetteries_data = {}
        current_dechetterie = None
        months_order = ['JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN',
                       'JUILLET', 'AOUT', 'SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DECEMBRE']
        
        # Trouver la ligne de début des données (après le titre)
        start_row = None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if any(cell and 'COLLECTES' in str(cell).upper() for cell in row if cell):
                start_row = row_idx + 2  # Après le titre et la ligne vide
                break
        
        if start_row is None:
            start_row = 3
        
        # Liste des noms de déchetteries connus pour éviter les faux positifs
        known_dechetteries = ['Pépinière', 'Sanssac', 'St Germain', 'Polignac', 
                             'Yssingeaux', 'Bas-en-basset', 'Monistrol']
        known_dechetteries_upper = [d.upper() for d in known_dechetteries]
        
        # Mapping des noms en majuscules vers les noms normalisés
        dechetterie_name_mapping = {
            'PEPINIERE': 'Pépinière',
            'PÉPINIÈRE': 'Pépinière',
            'SANSSAC': 'Sanssac',
            'ST GERMAIN': 'St Germain',
            'SAINT-GERMAIN': 'St Germain',
            'POLIGNAC': 'Polignac',
            'YSSINGEAUX': 'Yssingeaux',
            'BAS-EN-BASSET': 'Bas-en-basset',
            'MONISTROL': 'Monistrol'
        }
        
        # Parser ligne par ligne
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            if not row or not row[0]:
                continue
                
            first_cell = str(row[0]).strip() if row[0] else ''
            first_cell_upper = first_cell.upper()
            
            # Détecter un nom de déchetterie (header en majuscules, pas un mois, pas un total)
            # Vérifier que c'est bien un nom de déchetterie connu
            is_dechetterie_header = (
                first_cell and 
                first_cell_upper not in months_order and 
                'TOTAL' not in first_cell_upper and 
                'POURCENTAGE' not in first_cell_upper and
                'COLLECTES' not in first_cell_upper and
                len(first_cell) > 2 and
                not first_cell.startswith('COL_') and
                not first_cell.strip() == '' and
                # Vérifier que c'est un nom de déchetterie connu (en majuscules ou normalisé)
                (first_cell_upper in known_dechetteries_upper or 
                 first_cell_upper in dechetterie_name_mapping or
                 first_cell in known_dechetteries)
            )
            
            if is_dechetterie_header:
                # C'est un header de déchetterie
                # Normaliser le nom AVANT de créer l'entrée
                normalized_name = dechetterie_name_mapping.get(first_cell_upper, first_cell)
                
                # Si le nom normalisé n'est pas dans la liste connue, ignorer
                if normalized_name not in known_dechetteries and first_cell not in known_dechetteries:
                    # Ce n'est probablement pas une déchetterie, ignorer
                    continue
                
                # Si on avait une déchetterie précédente différente, on la finalise
                if current_dechetterie and current_dechetterie != normalized_name and current_dechetterie in dechetteries_data:
                    # Calculer les totaux si pas déjà fait
                    if not dechetteries_data[current_dechetterie]['total'] or not any(v for v in dechetteries_data[current_dechetterie]['total'].values() if isinstance(v, (int, float)) and v > 0):
                        totals_by_category = {col: 0 for col in category_columns}
                        grand_total = 0
                        for month_data in dechetteries_data[current_dechetterie]['months'].values():
                            for col in category_columns:
                                totals_by_category[col] += month_data.get(col, 0)
                            grand_total += month_data.get('TOTAL', 0)
                        totals_by_category['TOTAL'] = grand_total
                        dechetteries_data[current_dechetterie]['total'] = totals_by_category
                
                # Utiliser le nom normalisé
                current_dechetterie = normalized_name
                
                # Créer l'entrée seulement si elle n'existe pas déjà
                if current_dechetterie not in dechetteries_data:
                    dechetteries_data[current_dechetterie] = {
                        'months': {},
                        'total': {},
                        'categories': {col: [] for col in category_columns}  # Pour stocker les valeurs par catégorie
                    }
            elif first_cell_upper == 'TOTAL' and current_dechetterie:
                # C'est la ligne de total pour la déchetterie actuelle
                total_data = {}
                total_value = 0
                
                # Lire les valeurs de la ligne Total en utilisant les indices de colonnes détectés
                for col_name in category_columns:
                    col_idx = column_indices.get(col_name)
                    if col_idx and col_idx <= len(row):
                        val = row[col_idx - 1]  # -1 car row est 0-indexed mais col_idx est 1-indexed
                        # Gérer les formules Excel (elles sont déjà calculées avec data_only=True)
                        if isinstance(val, str) and val.startswith('='):
                            # C'est une formule, on ne peut pas la lire directement
                            total_data[col_name] = 0
                        elif pd.notna(val) and isinstance(val, (int, float)):
                            total_data[col_name] = float(val)
                            total_value += float(val)
                        else:
                            total_data[col_name] = 0
                    else:
                        total_data[col_name] = 0
                
                # Chercher la colonne TOTAL (généralement après toutes les catégories)
                # On cherche une colonne qui contient "TOTAL" ou qui est après toutes les catégories
                total_col_idx = None
                max_cat_col = max(column_indices.values()) if column_indices else len(category_columns) + 1
                for col_idx in range(max_cat_col + 1, min(max_cat_col + 5, len(row) + 1)):
                    if col_idx <= len(row):
                        cell_val = str(row[col_idx - 1]).upper() if row[col_idx - 1] else ''
                        if 'TOTAL' in cell_val:
                            total_col_idx = col_idx
                            break
                
                if total_col_idx and total_col_idx <= len(row):
                    total_val = row[total_col_idx - 1]
                    if isinstance(total_val, str) and total_val.startswith('='):
                        # Formule, on ne peut pas la lire
                        total_value = 0
                    elif pd.notna(total_val) and isinstance(total_val, (int, float)):
                        total_value = float(total_val)
                
                # Si total_value est 0, calculer depuis les données mensuelles
                if total_value == 0:
                    for month_data in dechetteries_data[current_dechetterie]['months'].values():
                        for col_name in category_columns:
                            total_data[col_name] += month_data.get(col_name, 0)
                        total_value += month_data.get('TOTAL', 0)
                
                # DECHETS ULTIMES total (colonne dédiée)
                if ultimes_col_idx and ultimes_col_idx <= len(row):
                    ult_val = row[ultimes_col_idx - 1]
                    if pd.notna(ult_val) and isinstance(ult_val, (int, float)):
                        total_data['DECHETS ULTIMES'] = float(ult_val)
                    else:
                        total_data['DECHETS ULTIMES'] = 0
                else:
                    total_data['DECHETS ULTIMES'] = 0

                total_data['TOTAL'] = total_value
                dechetteries_data[current_dechetterie]['total'] = total_data
                # Ne pas reset current_dechetterie ici, car il peut y avoir une ligne vide après
            elif first_cell_upper in months_order and current_dechetterie:
                # C'est une ligne de données mensuelles
                month = first_cell_upper  # Utiliser la version en majuscules pour la cohérence
                month_data = {}
                total = 0
                
                # S'assurer que la déchetterie existe
                if current_dechetterie not in dechetteries_data:
                    dechetteries_data[current_dechetterie] = {
                        'months': {},
                        'total': {},
                        'categories': {col: [] for col in category_columns}
                    }
                
                # Lire les valeurs mensuelles en utilisant les indices de colonnes détectés
                for col_name in category_columns:
                    col_idx = column_indices.get(col_name)
                    if col_idx and col_idx <= len(row):
                        val = row[col_idx - 1]  # -1 car row est 0-indexed mais col_idx est 1-indexed
                        if pd.notna(val) and isinstance(val, (int, float)):
                            month_data[col_name] = float(val)
                            total += float(val)
                            # Ajouter à la liste des valeurs pour cette catégorie
                            dechetteries_data[current_dechetterie]['categories'][col_name].append(float(val))
                        else:
                            month_data[col_name] = 0
                    else:
                        month_data[col_name] = 0
                
                # Chercher la colonne TOTAL
                max_cat_col = max(column_indices.values()) if column_indices else len(category_columns) + 1
                total_col_idx = None
                for col_idx in range(max_cat_col + 1, min(max_cat_col + 5, len(row) + 1)):
                    if col_idx <= len(row):
                        cell_val = str(row[col_idx - 1]).upper() if row[col_idx - 1] else ''
                        if 'TOTAL' in cell_val or (isinstance(row[col_idx - 1], (int, float)) and col_idx == max_cat_col + 1):
                            total_col_idx = col_idx
                            break
                
                if total_col_idx and total_col_idx <= len(row):
                    total_val = row[total_col_idx - 1]
                    if pd.notna(total_val) and isinstance(total_val, (int, float)):
                        total = float(total_val)
                
                # DECHETS ULTIMES mensuel
                if ultimes_col_idx and ultimes_col_idx <= len(row):
                    ult_val = row[ultimes_col_idx - 1]
                    if pd.notna(ult_val) and isinstance(ult_val, (int, float)):
                        month_data['DECHETS ULTIMES'] = float(ult_val)
                    else:
                        month_data['DECHETS ULTIMES'] = 0
                else:
                    month_data['DECHETS ULTIMES'] = 0

                month_data['TOTAL'] = total
                dechetteries_data[current_dechetterie]['months'][month] = month_data
        
        wb.close()
        
        # Finaliser les totaux pour toutes les déchetteries et nettoyer les doublons
        # D'abord, calculer les totaux manquants
        for dech, data in dechetteries_data.items():
            if not data['total'] or not any(v for v in data['total'].values() if isinstance(v, (int, float)) and v > 0):
                # Calculer les totaux depuis les données mensuelles
                totals_by_category = {col: 0 for col in category_columns}
                grand_total = 0
                
                for month, month_data in data['months'].items():
                    for col in category_columns:
                        totals_by_category[col] += month_data.get(col, 0)
                    grand_total += month_data.get('TOTAL', 0)
                
                totals_by_category['DECHETS ULTIMES'] = sum(
                    month_data.get('DECHETS ULTIMES', 0) for month_data in data['months'].values()
                )
                totals_by_category['TOTAL'] = grand_total
                data['total'] = totals_by_category
        
        # Nettoyer les doublons potentiels (même nom en majuscules/minuscules)
        cleaned_dechetteries = {}
        seen_normalized = set()
        
        for dech_name, data in dechetteries_data.items():
            # Normaliser le nom pour la comparaison
            normalized = dechetterie_name_mapping.get(dech_name.upper(), dech_name)
            
            if normalized not in seen_normalized:
                # Première occurrence, on la garde
                cleaned_dechetteries[normalized] = data
                seen_normalized.add(normalized)
            else:
                # Doublon détecté, fusionner les données
                existing_data = cleaned_dechetteries[normalized]
                # Fusionner les mois
                for month, month_data in data['months'].items():
                    if month not in existing_data['months']:
                        existing_data['months'][month] = month_data
                    else:
                        # Fusionner les valeurs
                        for col in category_columns:
                            existing_data['months'][month][col] = existing_data['months'][month].get(col, 0) + month_data.get(col, 0)
                        existing_data['months'][month]['TOTAL'] = existing_data['months'][month].get('TOTAL', 0) + month_data.get('TOTAL', 0)
                        existing_data['months'][month]['DECHETS ULTIMES'] = (
                            existing_data['months'][month].get('DECHETS ULTIMES', 0)
                            + month_data.get('DECHETS ULTIMES', 0)
                        )
                
                # Recalculer les totaux après fusion
                totals_by_category = {col: 0 for col in category_columns}
                grand_total = 0
                for month_data in existing_data['months'].values():
                    for col in category_columns:
                        totals_by_category[col] += month_data.get(col, 0)
                    grand_total += month_data.get('TOTAL', 0)
                    totals_by_category['DECHETS ULTIMES'] += month_data.get('DECHETS ULTIMES', 0)
                totals_by_category['TOTAL'] = grand_total
                existing_data['total'] = totals_by_category
        
        dechetteries_data = cleaned_dechetteries
        
        # Calculer les totaux globaux
        global_totals = {col: 0 for col in category_columns}
        global_totals['DECHETS ULTIMES'] = 0
        global_totals['TOTAL'] = 0
        
        for dech, data in dechetteries_data.items():
            for col in category_columns:
                global_totals[col] += data['total'].get(col, 0)
            global_totals['TOTAL'] += data['total'].get('TOTAL', 0)
            global_totals['DECHETS ULTIMES'] += data['total'].get('DECHETS ULTIMES', 0)
        
        # Préparer les données pour les graphiques
        stats = {
            'dechetteries': dechetteries_data,
            'global_totals': global_totals,
            'category_columns': category_columns,
            'final_fluxes': final_fluxes,
            'months_order': months_order,
            'num_dechetteries': len(dechetteries_data),
            'num_months': len([m for m in months_order if any(
                dech_data['months'].get(m) for dech_data in dechetteries_data.values()
            )]),
            'dataset_year': dataset_year,
            'date_start': date_start.strftime('%Y-%m-%d') if date_start else None,
            'date_end': date_end.strftime('%Y-%m-%d') if date_end else None,
            'date_range': date_range_label,
        }
        
        return {
            'success': True,
            'stats': stats,
            'error': None
        }
        
    except Exception as e:
        return {
            'success': False,
            'stats': None,
            'error': str(e)
        }
